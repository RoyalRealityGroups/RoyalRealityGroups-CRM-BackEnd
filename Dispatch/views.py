import logging
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum, Count
from rest_framework import generics, permissions, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from django_filters import FilterSet, CharFilter, DateFilter, NumberFilter, UUIDFilter, ChoiceFilter
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, timedelta, date
from decimal import Decimal

from Core.Core.permissions.permissions import GetPermission

from .models import DispatchPlan, DispatchItem, DispatchOrderItem
from .attachment_models import DispatchPlanAttachment
from .serializers import (
    DispatchPlanSerializer,
    DispatchPlanListSerializer,
    DispatchItemSerializer,
    DispatchPlanStatusCountSerializer,
    SalesOrderForDispatchSerializer,
    DispatchPlanningReportSerializer,
)
from Sales.models import SalesOrder
from Masters.models import RouteCoverage, Location
from utils import apply_company_location_filter

logger = logging.getLogger(__name__)


def _set_draft_authorization(instance):
    """Keep draft records out of authorization workflow."""
    instance.authorized_status = 0
    instance.current_authorized_status = 0
    instance.authorized_level = 0
    instance.current_authorized_level = 0
    instance.authorized_by_type = None
    instance.authorized_by_identifier = None
    instance.authorized_on = None
    instance.current_authorized_by_type = None
    instance.current_authorized_by_identifier = None
    instance.current_authorized_on = None
    instance.save(update_fields=[
        'authorized_status', 'current_authorized_status',
        'authorized_level', 'current_authorized_level',
        'authorized_by_type', 'authorized_by_identifier', 'authorized_on',
        'current_authorized_by_type', 'current_authorized_by_identifier', 'current_authorized_on',
    ])


def _sync_dispatch_status_from_authorization(instance):
    if instance.status == 'DRAFT':
        return

    target_status = 'CONFIRMED' if instance.authorized_status == 2 else 'PENDING'
    if instance.status != target_status:
        DispatchPlan.objects.filter(pk=instance.pk).update(status=target_status)
        instance.status = target_status


class DispatchPlanFilter(FilterSet):
    """Filter for Dispatch Plans"""
    dispatch_number = CharFilter(field_name='dispatch_number', lookup_expr='icontains')
    status = CharFilter(field_name='status')
    dispatch_date_from = DateFilter(field_name='dispatch_date', lookup_expr='gte')
    dispatch_date_to = DateFilter(field_name='dispatch_date', lookup_expr='lte')
    authorized_status = NumberFilter(field_name='authorized_status')
    
    class Meta:
        model = DispatchPlan
        fields = ['dispatch_number', 'status', 'authorized_status', 'dispatch_date_from', 'dispatch_date_to']





class DispatchPlanList(generics.ListCreateAPIView):
    """List and create dispatch plans"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = DispatchPlanFilter
    search_fields = [
        'dispatch_number', 
        'driver_name',
        'driver_phone',
        'vehicle_number',
        'remarks'
    ]
    ordering_fields = ['dispatch_date', 'dispatch_number', 'total_value', 'authorized_status', 'authorized_on', 'created_on']
    ordering = ['-dispatch_date', '-created_on']
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DispatchPlanSerializer
        return DispatchPlanListSerializer
    
    def get_queryset(self):
        queryset = DispatchPlan.objects.filter(
            is_deleted=False
        ).select_related('location').prefetch_related('items')
        return apply_company_location_filter(
            queryset, self.request.user,
            company_field=None, location_field='location'
        )

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        
        # Get location to generate dispatch number (only if location provided)
        location_id = data.get('location')
        if location_id:
            try:
                location = Location.objects.get(id=location_id)
                data['dispatch_number'] = DispatchPlan.generate_dispatch_number(location.code)
            except Location.DoesNotExist:
                return Response({'error': 'Invalid location'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Fallback to old format for backward compatibility
            data['dispatch_number'] = f"DP-{DispatchPlan.objects.count() + 1}"
        
        items_data = data.pop('items', [])
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        dispatch_plan = serializer.save(
            created_by_type='User',
            created_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
        )
        requested_status = str(data.get('status', '')).upper()
        if requested_status == 'DRAFT' or (dispatch_plan.status == 'DRAFT' and requested_status == ''):
            _set_draft_authorization(dispatch_plan)
        else:
            _sync_dispatch_status_from_authorization(dispatch_plan)
        
        # Create dispatch items
        for item_data in items_data:
            sales_order = SalesOrder.objects.only('id', 'company_id').get(id=item_data['sales_order'])
            dispatch_item = DispatchItem.objects.create(
                dispatch_plan=dispatch_plan,
                sales_order_id=item_data['sales_order'],
                company_id=sales_order.company_id,
                quantity_ordered=item_data['quantity_ordered'],
                quantity_dispatched=item_data['quantity_dispatched'],
                delivery_sequence=item_data.get('delivery_sequence', 1),
                loading_sequence=item_data.get('loading_sequence', 1),
                unloading_sequence=item_data.get('unloading_sequence'),
                estimated_delivery_time=item_data.get('estimated_delivery_time'),
                delivery_notes=item_data.get('delivery_notes', ''),
                created_by_type='User',
                created_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
            )
            
            # Create DispatchOrderItem if sales_order_item is provided (item-level dispatch)
            if 'sales_order_item' in item_data and item_data['sales_order_item']:
                DispatchOrderItem.objects.create(
                    dispatch_item=dispatch_item,
                    sales_order_item_id=item_data['sales_order_item'],
                    company_id=dispatch_item.company_id,
                    quantity_ordered=item_data['quantity_ordered'],
                    quantity_dispatched=item_data['quantity_dispatched'],
                    notes=item_data.get('delivery_notes', ''),
                    created_by_type='User',
                    created_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
                )
        
        # Calculate totals
        dispatch_plan.calculate_totals()
        
        # Update sales order dispatch status
        for item_data in items_data:
            sales_order = SalesOrder.objects.get(id=item_data['sales_order'])
            sales_order.update_dispatch_status()
        
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class DispatchPlanDetail(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a dispatch plan"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [JSONParser]
    serializer_class = DispatchPlanSerializer
    
    def get_queryset(self):
        queryset = DispatchPlan.objects.filter(
            is_deleted=False
        ).select_related('location', 'route').prefetch_related(
            'items',
            'items__sales_order',
            'items__sales_order__shipping_state',
            'items__sales_order__shipping_city',
            'items__order_items',
            'items__order_items__sales_order_item',
            'attachments'
        )
        return apply_company_location_filter(
            queryset, self.request.user,
            company_field=None, location_field='location'
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Only allow editing until authorization is approved
        if instance.authorized_status == 2:  # APPROVED
            return Response(
                {'error': 'Cannot edit dispatch plan that has been approved. Only pending or rejected dispatch plans can be edited.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if dispatch plan has confirmed invoices
        has_confirmed_invoices = instance.invoices.filter(
            status__in=['CONFIRMED', 'PAID', 'PARTIALLY_PAID']
        ).exists()
        
        if has_confirmed_invoices:
            return Response(
                {'error': 'Cannot edit dispatch plan with confirmed invoices'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = request.data.copy()
        items_data = data.pop('items', [])
        
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        requested_status = str(data.get('status', '')).upper()
        dispatch_plan = serializer.save(
            modified_by_type='User',
            modified_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
        )
        if requested_status == 'DRAFT':
            if dispatch_plan.status != 'DRAFT':
                dispatch_plan.status = 'DRAFT'
                dispatch_plan.save(update_fields=['status'])
            _set_draft_authorization(dispatch_plan)
        else:
            _sync_dispatch_status_from_authorization(dispatch_plan)
        
        # Update dispatch items
        if items_data:
            # Check if any existing items have invoices before deleting
            for existing_item in dispatch_plan.items.all():
                if existing_item.get_invoiced_quantity() > 0:
                    return Response(
                        {'error': f'Cannot modify dispatch - items already invoiced'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Delete existing items
            dispatch_plan.items.all().delete()
            
            # Create new items
            for item_data in items_data:
                sales_order = SalesOrder.objects.only('id', 'company_id').get(id=item_data['sales_order'])
                dispatch_item = DispatchItem.objects.create(
                    dispatch_plan=dispatch_plan,
                    sales_order_id=item_data['sales_order'],
                    company_id=sales_order.company_id,
                    quantity_ordered=item_data['quantity_ordered'],
                    quantity_dispatched=item_data['quantity_dispatched'],
                    delivery_sequence=item_data.get('delivery_sequence', 1),
                    loading_sequence=item_data.get('loading_sequence', 1),
                    unloading_sequence=item_data.get('unloading_sequence'),
                    estimated_delivery_time=item_data.get('estimated_delivery_time'),
                    delivery_notes=item_data.get('delivery_notes', ''),
                    modified_by_type='User',
                    modified_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
                )
                
                # Create DispatchOrderItem if sales_order_item is provided (item-level dispatch)
                if 'sales_order_item' in item_data and item_data['sales_order_item']:
                    DispatchOrderItem.objects.create(
                        dispatch_item=dispatch_item,
                        sales_order_item_id=item_data['sales_order_item'],
                        company_id=dispatch_item.company_id,
                        quantity_ordered=item_data['quantity_ordered'],
                        quantity_dispatched=item_data['quantity_dispatched'],
                        notes=item_data.get('delivery_notes', ''),
                        created_by_type='User',
                        created_by_identifier=str(request.user.id) if request.user.is_authenticated else 'Anonymous'
                    )
            
            # Recalculate totals
            dispatch_plan.calculate_totals()
            
            # Update sales order dispatch status
            for item_data in items_data:
                sales_order = SalesOrder.objects.get(id=item_data['sales_order'])
                sales_order.update_dispatch_status()
        
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        """Only allow deletion of DRAFT plans without invoices"""
        instance = self.get_object()
        if instance.status not in ['DRAFT']:
            return Response(
                {'error': 'Only DRAFT dispatch plans can be deleted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if any items have invoices
        has_invoices = instance.invoices.filter(
            status__in=['CONFIRMED', 'PAID', 'PARTIALLY_PAID']
        ).exists()
        
        if has_invoices:
            return Response(
                {'error': 'Cannot delete dispatch plan with confirmed invoices'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().destroy(request, *args, **kwargs)


class AvailableOrdersForDispatch(generics.ListCreateAPIView):
    """Get sales orders available for dispatch"""
    permission_classes = [permissions.AllowAny]
    serializer_class = SalesOrderForDispatchSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = [
        'order_number', 
        'retailer__name', 
        'retailer__code',
        'distributor__name', 
        'distributor__code',
        'superstockist__name',
        'superstockist__code',
        'shipping_city__name',
        'shipping_state__name'
    ]
    
    def get_queryset(self):
        queryset = SalesOrder.objects.filter(
            status__in=['CONFIRMED', 'PARTIALLY_DISPATCHED', 'PARTIALLY_INVOICED'],
            authorized_status=2,  # Only APPROVED orders
            is_deleted=False
        ).select_related(
            'company', 'shipping_state', 'shipping_city', 'shipping_area',
            'retailer', 'distributor', 'superstockist'
        ).prefetch_related('items', 'dispatch_items')
        return apply_company_location_filter(
            queryset, self.request.user,
            company_field='company', location_field=None
        )
    
    def post(self, request, *args, **kwargs):
        """Handle POST request with filters in body"""
        data = request.data
        customer_type = data.get('customer_type')
        state_id = data.get('state')
        city_id = data.get('city')
        area_id = data.get('area')
        route_id = data.get('route_id')
        customers = data.get('customer')
        dispatch_plan_id = data.get('dispatch_plan_id')  # For edit mode
        
        queryset = self.get_queryset()

        # For create mode, require either route or specific customer to avoid loading all pending orders.
        if not route_id and not dispatch_plan_id and not customers:
            queryset = queryset.none()

        if customer_type:
            queryset = queryset.filter(customer_type=customer_type)
        if state_id:
            queryset = queryset.filter(shipping_state_id=state_id)
        if city_id:
            queryset = queryset.filter(shipping_city_id=city_id)
        if area_id:
            queryset = queryset.filter(shipping_area_id=area_id)
        if route_id:
            route_coverages = RouteCoverage.objects.filter(route_id=route_id)
            route_area_ids = list(route_coverages.values_list('area_id', flat=True))
            route_city_ids = list(route_coverages.values_list('city_id', flat=True).distinct())
            route_state_ids = list(route_coverages.values_list('state_id', flat=True).distinct())

            if not route_area_ids and not route_city_ids and not route_state_ids:
                queryset = queryset.none()
            else:
                route_filter = Q()
                if route_area_ids:
                    route_filter |= Q(shipping_area_id__in=route_area_ids)
                if route_city_ids:
                    route_filter |= Q(shipping_area__isnull=True, shipping_city_id__in=route_city_ids)
                if route_state_ids:
                    route_filter |= Q(shipping_area__isnull=True, shipping_city__isnull=True, shipping_state_id__in=route_state_ids)
                queryset = queryset.filter(route_filter)
        if customers:
            # Customer-first flow: filter by selected customer id regardless of type,
            # then optionally constrain by explicit customer_type if provided.
            queryset = queryset.filter(
                Q(retailer_id=customers) |
                Q(distributor_id=customers) |
                Q(superstockist_id=customers)
            )
            if customer_type:
                queryset = queryset.filter(customer_type=customer_type)
        
        # Filter out fully dispatched orders (except in edit mode where we need to show selected orders)
        available_orders = []
        for order in queryset:
            total_qty = sum(item.quantity for item in order.items.all())
            
            # Exclude CANCELLED dispatches and REJECTED dispatches (authorized_status=3)
            excluded_dispatches_q = Q(dispatch_plan__status='CANCELLED') | Q(dispatch_plan__authorized_status=3)
            if dispatch_plan_id:
                # Edit mode: exclude dispatches from other plans, include current plan's dispatches
                dispatched_qty = sum(
                    dispatch_item.quantity_dispatched 
                    for dispatch_item in order.dispatch_items.exclude(
                        excluded_dispatches_q
                    ).exclude(dispatch_plan_id=dispatch_plan_id)
                )
            else:
                # Create mode: exclude all dispatches
                dispatched_qty = sum(
                    dispatch_item.quantity_dispatched 
                    for dispatch_item in order.dispatch_items.exclude(excluded_dispatches_q)
                )
            
            remaining_qty = total_qty - dispatched_qty
            # In edit mode, include orders even with 0 remaining if they're part of current dispatch
            if dispatch_plan_id:
                is_in_current_dispatch = order.dispatch_items.filter(dispatch_plan_id=dispatch_plan_id).exists()
                if remaining_qty > 0 or is_in_current_dispatch:
                    available_orders.append(order)
            elif remaining_qty > 0:
                available_orders.append(order)
        
        # Apply pagination to filtered results
        page = self.paginate_queryset(available_orders)
        if page is not None:
            serializer = self.get_serializer(page, many=True, context={
                'request': request, 
                'show_items': True,
                'dispatch_plan_id': dispatch_plan_id
            })
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(available_orders, many=True, context={
            'request': request, 
            'show_items': True,
            'dispatch_plan_id': dispatch_plan_id
        })
        return Response({'results': serializer.data, 'count': len(available_orders)})


class GenerateDispatchNumber(APIView):
    """Generate new dispatch number"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        location_id = request.query_params.get('location')
        if not location_id:
            return Response({'error': 'Location parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            location = Location.objects.get(id=location_id)
            dispatch_number = DispatchPlan.generate_dispatch_number(location.code)
            return Response({'dispatch_number': dispatch_number})
        except Location.DoesNotExist:
            return Response({'error': 'Invalid location'}, status=status.HTTP_400_BAD_REQUEST)


class DispatchPlanAttachmentList(APIView):
    """List and create attachments for a dispatch plan"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request, pk):
        attachments = DispatchPlanAttachment.objects.filter(dispatch_plan_id=pk)
        data = [{
            'id': att.id,
            'attachment_type': att.attachment_type,
            'file_url': request.build_absolute_uri(att.file.url) if att.file else None,
            'original_filename': att.original_filename,
            'description': att.description,
        } for att in attachments]
        return Response(data)

    def post(self, request, pk):
        plan = get_object_or_404(DispatchPlan, pk=pk, is_deleted=False)
        type_key_map = {
            'vehicle_rc': 'VEHICLE_RC',
            'vehicle_insurance': 'VEHICLE_INSURANCE',
            'vehicle_permit': 'VEHICLE_PERMIT',
            'vehicle_pollution': 'VEHICLE_POLLUTION',
            'driver_license': 'DRIVER_LICENSE',
            'other': 'OTHER',
        }

        created = []
        for key, att_type in type_key_map.items():
            files = request.FILES.getlist(key)
            if not files:
                continue
            # For typed attachments (non-OTHER), replace existing
            if att_type != 'OTHER':
                DispatchPlanAttachment.objects.filter(dispatch_plan=plan, attachment_type=att_type).delete()
                files = files[:1]
            for file in files:
                att = DispatchPlanAttachment.objects.create(
                    dispatch_plan=plan,
                    attachment_type=att_type,
                    file=file,
                    original_filename=file.name,
                )
                created.append({
                    'id': att.id,
                    'attachment_type': att.attachment_type,
                    'file_url': request.build_absolute_uri(att.file.url),
                    'original_filename': att.original_filename,
                })

        if not created:
            return Response({'error': 'No files provided'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(created, status=status.HTTP_201_CREATED)


class DispatchPlanAttachmentDetail(APIView):
    """Delete a single dispatch plan attachment"""
    permission_classes = [permissions.AllowAny]

    def delete(self, request, pk, att_pk):
        att = get_object_or_404(DispatchPlanAttachment, pk=att_pk, dispatch_plan_id=pk)
        att.file.delete(save=False)
        att.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)



class DispatchPlanStatusCountView(generics.ListAPIView):
    permission_classes = [GetPermission('Dispatch.view_dispatchplan')]
    serializer_class = DispatchPlanStatusCountSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = DispatchPlanFilter

    def get_queryset(self):
        queryset = DispatchPlan.objects.filter(is_deleted=False)
        queryset = apply_company_location_filter(
            queryset,
            self.request.user,
            company_field=None,
            location_field='location'
        )
        return queryset

    def list(self, request, *args, **kwargs):
        # Get filtered queryset (applies all filter backends and filterset)
        queryset = self.filter_queryset(self.get_queryset())
        
        # Get counts by status
        status_counts = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Build response with all statuses
        all_statuses = {
            'DRAFT': 0,
            'PENDING': 0,
            'CONFIRMED': 0,
            'DELIVERED': 0,
            'CANCELLED': 0,
        }
        
        # Populate with actual counts
        for item in status_counts:
            if item['status'] in all_statuses:
                all_statuses[item['status']] = item['count']
        
        # Add total count
        all_statuses['total'] = sum(all_statuses.values())
        
        # Serialize the response
        serializer = self.get_serializer(all_statuses)
        return Response(serializer.data, status=status.HTTP_200_OK)


# ============================================================
# Dispatch Planning Report Views
# ============================================================

def _get_report_queryset(user):
    """Shared queryset builder for report view and export view."""
    queryset = DispatchPlan.objects.filter(
        is_deleted=False
    ).select_related(
        'location',
        'route',
    ).prefetch_related(
        'items',
        'items__sales_order',
        'items__sales_order__retailer',
        'items__sales_order__distributor',
        'items__sales_order__distributor__agent',
        'items__sales_order__superstockist',
        'items__sales_order__billing_state',
        'items__sales_order__billing_city',
        'items__sales_order__billing_area',
        'items__order_items',
        'items__order_items__sales_order_item',
        'items__order_items__sales_order_item__item',
    )
    return apply_company_location_filter(
        queryset, user,
        company_field=None, location_field='location'
    )


def _convert_to_product_rows(dispatch_plans):
    """
    Convert dispatch plans to product-wise rows.
    Each product in each order in each dispatch plan becomes a separate row.
    """
    product_rows = []

    for dispatch_plan in dispatch_plans:
        location_name = dispatch_plan.location.name if dispatch_plan.location else None
        location_code = dispatch_plan.location.code if dispatch_plan.location else None
        route_name = dispatch_plan.route.name if dispatch_plan.route else None
        route_code = dispatch_plan.route.code if dispatch_plan.route else None

        for dispatch_item in dispatch_plan.items.all():
            sales_order = dispatch_item.sales_order
            billing_state = sales_order.billing_state
            billing_city = sales_order.billing_city
            billing_area = sales_order.billing_area

            state = billing_state.name if billing_state else ''
            state_id = str(billing_state.id) if billing_state else None
            city = billing_city.name if billing_city else ''
            city_id = str(billing_city.id) if billing_city else None
            area = billing_area.name if billing_area else None
            area_id = str(billing_area.id) if billing_area else None

            # Shared fields for every row produced from this dispatch_item
            common = {
                'dispatch_plan_id': str(dispatch_plan.id),
                'dispatch_number': dispatch_plan.dispatch_number,
                'dispatch_date': dispatch_plan.dispatch_date,
                'planned_dispatch_date': dispatch_plan.planned_dispatch_date,
                'sales_order_id': str(sales_order.id),
                'sales_order_number': sales_order.order_number,
                'sales_order_date': sales_order.order_date,
                'customer_name': sales_order.get_customer_name(),
                'customer_type': sales_order.get_customer_type_display(),
                'customer_type_code': sales_order.customer_type,
                'location_name': location_name,
                'location_code': location_code,
                'route_name': route_name,
                'route_code': route_code,
                'state': state,
                'state_id': state_id,
                'city': city,
                'city_id': city_id,
                'area': area,
                'area_id': area_id,
                'dispatch_status': dispatch_plan.get_status_display(),
                'dispatch_status_code': dispatch_plan.status,
                'item_status': dispatch_item.get_status_display(),
                'item_status_code': dispatch_item.status,
                'vehicle_number': dispatch_plan.vehicle_number,
                'driver_name': dispatch_plan.driver_name,
                'driver_phone': dispatch_plan.driver_phone,
                'delivery_sequence': dispatch_item.delivery_sequence,
                'estimated_delivery_time': dispatch_item.estimated_delivery_time,
                'actual_delivery_time': dispatch_item.actual_delivery_time,
                'authorization_status': sales_order.get_authorized_status_display() if sales_order.authorized_status else 'Pending',
                'authorization_status_code': sales_order.authorized_status or 1,
                'agent_name': sales_order.distributor.agent.name if sales_order.distributor and sales_order.distributor.agent else None,
            }

            order_items = dispatch_item.order_items.all()
            if order_items:
                # Item-level rows (DispatchOrderItem exists)
                for order_item in order_items:
                    product = order_item.sales_order_item.item
                    product_rows.append({
                        **common,
                        'dispatch_order_item_id': str(order_item.id),
                        'product_id': str(product.id),
                        'product_code': product.code,
                        'product_name': product.name,
                        'quantity_ordered': order_item.quantity_ordered,
                        'quantity_dispatched': order_item.quantity_dispatched,
                        # Use the DispatchOrderItem's own status, not the parent DispatchItem's
                        'item_status': order_item.get_status_display(),
                        'item_status_code': order_item.status,
                    })
            else:
                # Order-level fallback — no DispatchOrderItem children
                product_rows.append({
                    **common,
                    'dispatch_order_item_id': str(dispatch_item.id),
                    'product_id': None,
                    'product_code': '',
                    'product_name': f'(Order {sales_order.order_number})',
                    'quantity_ordered': dispatch_item.quantity_ordered,
                    'quantity_dispatched': dispatch_item.quantity_dispatched,
                })

    return product_rows


# Headers shared between Excel and CSV exports
_EXPORT_HEADERS = [
    'Dispatch Number', 'Dispatch Date', 'Sales Order', 'Order Date',
    'Customer', 'Customer Type', 'Product Code', 'Product Name',
    'Qty Ordered', 'Qty Dispatched', 'Location', 'Route',
    'State', 'City', 'Status', 'Vehicle',
]


def _iter_export_rows(queryset):
    """Yield one flat row tuple per product across all dispatch plans."""
    for dispatch_plan in queryset:
        location_name = dispatch_plan.location.name if dispatch_plan.location else ''
        route_name = dispatch_plan.route.name if dispatch_plan.route else ''

        for dispatch_item in dispatch_plan.items.all():
            sales_order = dispatch_item.sales_order
            so_state = sales_order.billing_state.name if sales_order.billing_state else ''
            so_city = sales_order.billing_city.name if sales_order.billing_city else ''

            order_items = dispatch_item.order_items.all()
            if order_items:
                for order_item in order_items:
                    product = order_item.sales_order_item.item
                    yield (
                        dispatch_plan.dispatch_number,
                        dispatch_plan.dispatch_date.strftime('%Y-%m-%d'),
                        sales_order.order_number,
                        sales_order.order_date.strftime('%Y-%m-%d'),
                        sales_order.get_customer_name(),
                        sales_order.get_customer_type_display(),
                        product.code,
                        product.name,
                        float(order_item.quantity_ordered),
                        float(order_item.quantity_dispatched),
                        location_name,
                        route_name,
                        so_state,
                        so_city,
                        dispatch_plan.get_status_display(),
                        dispatch_plan.vehicle_number or '',
                    )
            else:
                # Order-level fallback — no DispatchOrderItem children
                yield (
                    dispatch_plan.dispatch_number,
                    dispatch_plan.dispatch_date.strftime('%Y-%m-%d'),
                    sales_order.order_number,
                    sales_order.order_date.strftime('%Y-%m-%d'),
                    sales_order.get_customer_name(),
                    sales_order.get_customer_type_display(),
                    '',
                    f'(Order {sales_order.order_number})',
                    float(dispatch_item.quantity_ordered),
                    float(dispatch_item.quantity_dispatched),
                    location_name,
                    route_name,
                    so_state,
                    so_city,
                    dispatch_plan.get_status_display(),
                    dispatch_plan.vehicle_number or '',
                )


class DispatchPlanningReportFilter(FilterSet):
    """Advanced filter for Dispatch Planning Report."""
    from_date = DateFilter(field_name='dispatch_date', lookup_expr='gte', label='From Date')
    to_date = DateFilter(field_name='dispatch_date', lookup_expr='lte', label='To Date')
    date_preset = CharFilter(method='filter_date_preset', label='Date Preset')
    location = UUIDFilter(field_name='location_id', label='Location')
    route = UUIDFilter(field_name='route_id', label='Route')
    state = UUIDFilter(method='filter_state', label='State')
    city = UUIDFilter(method='filter_city', label='City')
    area = UUIDFilter(method='filter_area', label='Area')
    customer_type = ChoiceFilter(
        method='filter_customer_type',
        choices=[
            ('RETAILER', 'Retailer'),
            ('DISTRIBUTOR', 'Distributor'),
            ('SUPERSTOCKIST', 'Superstockist'),
        ],
        label='Customer Type'
    )
    customer_id = UUIDFilter(method='filter_customer', label='Customer ID')
    sales_order_number = CharFilter(method='filter_sales_order', label='Sales Order Number')
    dispatch_status = ChoiceFilter(
        field_name='status',
        choices=[
            ('DRAFT', 'Draft'),
            ('PENDING', 'Pending'),
            ('CONFIRMED', 'Confirmed'),
            ('DELIVERED', 'Delivered'),
            ('CANCELLED', 'Cancelled'),
        ],
        label='Dispatch Status'
    )
    authorization_status = NumberFilter(method='filter_authorization_status', label='Authorization Status')
    agent = UUIDFilter(method='filter_agent', label='Agent')

    class Meta:
        model = DispatchPlan
        fields = [
            'from_date', 'to_date', 'date_preset',
            'location', 'route',
            'state', 'city', 'area',
            'customer_type', 'customer_id',
            'sales_order_number', 'dispatch_status',
            'authorization_status'
        ]

    def filter_date_preset(self, queryset, name, value):
        today = date.today()
        presets = {
            'today': (today, today),
            'this_week': (today - timedelta(days=today.weekday()), today),
            'this_month': (today.replace(day=1), today),
            'this_year': (today.replace(month=1, day=1), today),
        }
        date_range = presets.get(value)
        if date_range:
            return queryset.filter(dispatch_date__gte=date_range[0], dispatch_date__lte=date_range[1])
        return queryset

    def filter_state(self, queryset, name, value):
        return queryset.filter(items__sales_order__billing_state_id=value).distinct()

    def filter_city(self, queryset, name, value):
        return queryset.filter(items__sales_order__billing_city_id=value).distinct()

    def filter_area(self, queryset, name, value):
        return queryset.filter(items__sales_order__billing_area_id=value).distinct()

    def filter_customer_type(self, queryset, name, value):
        return queryset.filter(items__sales_order__customer_type=value).distinct()

    def filter_customer(self, queryset, name, value):
        return queryset.filter(
            Q(items__sales_order__retailer_id=value) |
            Q(items__sales_order__distributor_id=value) |
            Q(items__sales_order__superstockist_id=value)
        ).distinct()

    def filter_sales_order(self, queryset, name, value):
        return queryset.filter(items__sales_order__order_number__icontains=value).distinct()

    def filter_authorization_status(self, queryset, name, value):
        return queryset.filter(items__sales_order__authorized_status=value).distinct()

    def filter_agent(self, queryset, name, value):
        return queryset.filter(items__sales_order__distributor__agent_id=value).distinct()


class DispatchPlanningReportView(generics.ListAPIView):
    """
    Product-wise Dispatch Planning Report API.
    Each row represents ONE product from a dispatch plan.

    GET /api/dispatch/reports/planning/
    """

    permission_classes = [GetPermission('Dispatch.view_dispatchplan')]
    serializer_class = DispatchPlanningReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = DispatchPlanningReportFilter

    search_fields = [
        'dispatch_number',
        'items__sales_order__order_number',
        'items__sales_order__retailer__name',
        'items__sales_order__distributor__name',
        'items__sales_order__superstockist__name',
        'location__name',
        'route__name',
        'vehicle_number',
    ]

    ordering_fields = [
        'dispatch_date',
        'dispatch_number',
        'planned_dispatch_date',
        'status',
    ]

    ordering = ['-dispatch_date', 'dispatch_number']

    def get_queryset(self):
        return _get_report_queryset(self.request.user)

    def _get_summary(self, queryset):
        """Compute summary stats using DB aggregation instead of Python loops."""
        agg = queryset.aggregate(
            total_dispatch_plans=Count('id'),
            total_orders=Count('items__sales_order', distinct=True),
            total_products=Count('items__order_items', distinct=True),
            total_quantity=Sum('items__order_items__quantity_dispatched'),
        )
        return {
            'total_dispatch_plans': agg['total_dispatch_plans'] or 0,
            'total_orders': agg['total_orders'] or 0,
            'total_products': agg['total_products'] or 0,
            'total_quantity': float(agg['total_quantity'] or 0),
        }

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        summary = self._get_summary(queryset)

        # Flatten to product rows BEFORE pagination so count/next/previous
        # reflect the actual number of product rows, not dispatch plans.
        product_rows = _convert_to_product_rows(queryset)

        page = self.paginate_queryset(product_rows)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['summary'] = summary
            return response

        serializer = self.get_serializer(product_rows, many=True)
        return Response({
            'results': serializer.data,
            'count': len(product_rows),
            'summary': summary,
        })


class DispatchPlanningReportExportView(APIView):
    """
    Export Dispatch Planning Report to Excel, CSV, or PDF.

    POST /api/dispatch/reports/planning/export/
    """

    permission_classes = [GetPermission('Dispatch.view_dispatchplan')]

    VALID_FORMATS = frozenset(['excel', 'csv', 'pdf'])

    def post(self, request):
        export_format = request.data.get('format', 'excel').lower()
        report_filters = request.data.get('filters', {})

        if export_format not in self.VALID_FORMATS:
            return Response(
                {'error': 'Invalid format. Must be excel, csv, or pdf'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = self._get_filtered_queryset(request, report_filters)
        handler = getattr(self, f'_export_{export_format}')
        return handler(queryset)

    def _get_filtered_queryset(self, request, report_filters):
        queryset = _get_report_queryset(request.user)
        filterset = DispatchPlanningReportFilter(report_filters, queryset=queryset, request=request)
        return filterset.qs

    def _export_excel(self, queryset):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dispatch Planning Report'

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(_EXPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, row_data in enumerate(_iter_export_rows(queryset), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=dispatch_planning_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    def _export_csv(self, queryset):
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=dispatch_planning_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        writer = csv.writer(response)
        writer.writerow(_EXPORT_HEADERS)

        for row in _iter_export_rows(queryset):
            writer.writerow(row)

        return response

    def _export_pdf(self, queryset):
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1,
        )

        elements.append(Paragraph('Dispatch Planning Report', title_style))
        elements.append(Spacer(1, 0.2 * inch))

        pdf_headers = [
            'Dispatch No.', 'Date', 'Order No.', 'Customer',
            'Product', 'Qty', 'Location', 'Route', 'Status',
        ]
        data = [pdf_headers]

        for full_row in _iter_export_rows(queryset):
            data.append([
                full_row[0],
                full_row[1],
                full_row[2],
                full_row[4][:15],
                full_row[7][:20],
                f"{full_row[9]:.3f}",
                full_row[10][:15],
                full_row[11][:15],
                full_row[14][:10],
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=dispatch_planning_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.write(pdf)
        return response
