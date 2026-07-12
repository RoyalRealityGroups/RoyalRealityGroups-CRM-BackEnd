from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters import FilterSet, CharFilter, DateFilter, NumberFilter, UUIDFilter, ChoiceFilter
from django_filters.rest_framework import DjangoFilterBackend
from decimal import Decimal
from datetime import datetime, timedelta, date
from django.db.models import F, Q, Count, Sum
from django.utils import timezone

from .models import Invoice, InvoiceItem, Payment
from .serializers import (
    InvoiceListSerializer,
    InvoiceDetailSerializer,
    DispatchPlanForInvoiceSerializer,
    SalesOrderForInvoiceSerializer,
    PendingInvoiceSerializer,
    InvoiceStatusCountSerializer,
    InvoiceReportSerializer,
)
from Sales.models import SalesOrder, SalesOrderItem
from Dispatch.models import DispatchPlan
from utils import apply_company_location_filter


def _set_draft_authorization(instance):
    """Keep draft records out of authorization workflow."""
    instance.authorized_status = 0
    instance.current_authorized_status = 0
    instance.authorized_level = 0
    instance.current_authorized_level = 0
    instance.authorized_by_type = None
    instance.authorized_by_identifier = None
    instance.authorized_on = None
    instance.current_authorized_by_type = None
    instance.current_authorized_by_identifier = None
    instance.current_authorized_on = None
    instance.save(update_fields=[
        'authorized_status', 'current_authorized_status',
        'authorized_level', 'current_authorized_level',
        'authorized_by_type', 'authorized_by_identifier', 'authorized_on',
        'current_authorized_by_type', 'current_authorized_by_identifier', 'current_authorized_on',
    ])


def _sync_invoice_status_from_authorization(instance):
    if instance.status == 'DRAFT':
        return

    target_status = 'CONFIRMED' if instance.authorized_status == 2 else 'PENDING'
    if instance.status != target_status:
        Invoice.objects.filter(pk=instance.pk).update(status=target_status)
        instance.status = target_status


class InvoiceFilter(FilterSet):
    """Filter for Invoices"""
    invoice_number = CharFilter(field_name='invoice_number', lookup_expr='icontains')
    status = CharFilter(field_name='status')
    invoice_date_from = DateFilter(field_name='invoice_date', lookup_expr='gte')
    invoice_date_to = DateFilter(field_name='invoice_date', lookup_expr='lte')
    location = CharFilter(field_name='location')
    source_type = CharFilter(field_name='source_type')
    authorized_status = NumberFilter(field_name='authorized_status')
    pod_status = CharFilter(field_name='pod_status')
    sales_order__retailer = CharFilter(field_name='sales_order__retailer')
    sales_order__distributor = CharFilter(field_name='sales_order__distributor')
    sales_order__superstockist = CharFilter(field_name='sales_order__superstockist')
    
    class Meta:
        model = Invoice
        fields = ['invoice_number', 'status', 'authorized_status', 'invoice_date_from', 'invoice_date_to', 'location', 'source_type', 'pod_status']


class InvoiceListCreateView(generics.ListAPIView):
    """List invoices"""
    permission_classes = [permissions.AllowAny]
    serializer_class = InvoiceListSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = InvoiceFilter
    search_fields = [
        'invoice_number',
        'code',
        'sales_order__order_number', 
        'sales_order__retailer__name', 
        'sales_order__retailer__code',
        'sales_order__distributor__name', 
        'sales_order__distributor__code',
        'sales_order__superstockist__name',
        'sales_order__superstockist__code',
        'remarks'
    ]
    ordering_fields = ['invoice_date', 'invoice_number', 'grand_total', 'authorized_status', 'authorized_on', 'created_on']
    ordering = ['-invoice_date', '-created_on']
    
    def get_queryset(self):
        queryset = Invoice.filtered_objects.get_qs(
            user=self.request.user
        ).select_related(
            'company', 'location', 'sales_order'
        ).prefetch_related('items')
        return apply_company_location_filter(
            queryset, self.request.user,
            company_field='company', location_field='location'
        )


class InvoiceDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete an invoice"""
    permission_classes = [permissions.AllowAny]
    serializer_class = InvoiceDetailSerializer
    
    def get_queryset(self):
        queryset = Invoice.filtered_objects.get_qs(
            user=self.request.user
        ).select_related(
            'company', 'location', 'sales_order', 'dispatch_plan',
            'sales_order__billing_state', 'sales_order__billing_city',
            'sales_order__shipping_state', 'sales_order__shipping_city'
        ).prefetch_related(
            'items', 'items__sales_order_item', 'items__sales_order_item__item',
            'payments'
        )
        return apply_company_location_filter(
            queryset, self.request.user,
            company_field='company', location_field='location'
        )

    def destroy(self, request, *args, **kwargs):
        """Only allow deletion of DRAFT invoices"""
        instance = self.get_object()
        if instance.status not in ['DRAFT']:
            return Response(
                {'error': 'Only DRAFT invoices can be deleted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update sales order status when invoice is deleted
        sales_order = instance.sales_order
        instance.delete()
        sales_order.update_invoice_status()
        
        return Response(status=status.HTTP_204_NO_CONTENT)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        requested_status = str(request.data.get('status', '')).upper()
        instance = self.get_object()
        if requested_status == 'DRAFT':
            if instance.status != 'DRAFT':
                instance.status = 'DRAFT'
                instance.save(update_fields=['status'])
            _set_draft_authorization(instance)
        else:
            _sync_invoice_status_from_authorization(instance)

        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=response.status_code)


class AvailableCustomersForInvoiceView(generics.ListAPIView):
    """Get customers who have sales orders available for invoicing"""
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'code']
    
    def get_queryset(self):
        from django.db.models import Q
        
        customer_type = self.request.query_params.get('customer_type')
        source_type = self.request.query_params.get('source_type', 'ORDER')
        
        if not customer_type:
            return []
        
        # Get sales orders available for invoicing
        if source_type == 'DISPATCH':
            from django.db.models import Sum, Q, OuterRef, Subquery, DecimalField
            from django.db.models.functions import Coalesce
            from Dispatch.models import DispatchOrderItem
            
            # Find dispatched quantities from approved dispatch plans
            dispatched_subquery = DispatchOrderItem.objects.filter(
                sales_order_item=OuterRef('pk'),
                dispatch_item__dispatch_plan__status__in=['CONFIRMED', 'DELIVERED'],
                dispatch_item__dispatch_plan__authorized_status=2  # Only approved dispatch plans
            ).values('sales_order_item').annotate(
                total=Coalesce(Sum('quantity_dispatched'), Decimal('0'))
            ).values('total')[:1]
            
            # Find already invoiced quantities from dispatch-based invoices
            invoiced_subquery = InvoiceItem.objects.filter(
                sales_order_item=OuterRef('pk'),
                invoice__source_type='DISPATCH',
                invoice__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID']
            ).values('sales_order_item').annotate(
                total=Coalesce(Sum('quantity'), Decimal('0'))
            ).values('total')[:1]
            
            # Find sales order items with dispatched qty > invoiced qty
            items_with_available = SalesOrderItem.filtered_objects.get_qs(
                user=self.request.user
            ).annotate(
                dispatched_qty=Coalesce(Subquery(dispatched_subquery, output_field=DecimalField()), Decimal('0')),
                invoiced_qty=Coalesce(Subquery(invoiced_subquery, output_field=DecimalField()), Decimal('0'))
            ).filter(
                dispatched_qty__gt=F('invoiced_qty')
            ).values_list('order', flat=True).distinct()
            
            # For DISPATCH source, rely on DispatchPlan authorization
            # Sales orders with dispatched items from approved dispatch plans are available
            available_orders = SalesOrder.filtered_objects.get_qs(
                user=self.request.user,
                id__in=items_with_available,
                is_deleted=False
            )
            available_orders = apply_company_location_filter(available_orders, self.request.user, company_field='company')
        else:
            # For ORDER source, get approved sales orders not yet invoiced
            available_orders = SalesOrder.filtered_objects.get_qs(
                user=self.request.user,
                authorized_status=2,
                is_deleted=False
            ).exclude(
                Q(invoices__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID'])
            )
            available_orders = apply_company_location_filter(available_orders, self.request.user, company_field='company')
        
        # Get unique customer IDs from available orders
        customer_field = f'{customer_type.lower()}'
        customer_ids = available_orders.filter(
            **{f'{customer_field}__isnull': False}
        ).values_list(customer_field, flat=True).distinct()
        
        # Get customer model based on type
        if customer_type == 'RETAILER':
            from Masters.models import Retailer
            return Retailer.objects.filter(id__in=customer_ids, is_deleted=False)
        elif customer_type == 'DISTRIBUTOR':
            from Masters.models import Distributor
            return Distributor.objects.filter(id__in=customer_ids, is_deleted=False)
        elif customer_type == 'SUPERSTOCKIST':
            from Masters.models import Superstockist
            return Superstockist.objects.filter(id__in=customer_ids, is_deleted=False)
        
        return []
    
    def get_serializer_class(self):
        customer_type = self.request.query_params.get('customer_type')
        
        # Return a simple serializer with id, name, code
        from rest_framework import serializers
        
        class CustomerSerializer(serializers.Serializer):
            id = serializers.UUIDField()
            name = serializers.CharField()
            code = serializers.CharField()
        
        return CustomerSerializer


class GenerateInvoiceFromDispatchView(APIView):
    """Generate invoice(s) from dispatch (based on sales order with dispatched quantities).
    Items are grouped by their company — one invoice per company."""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        sales_order_id = request.data.get('sales_order')
        location_id = request.data.get('location')
        invoice_date = request.data.get('invoice_date')
        due_date = request.data.get('due_date')
        invoice_number = request.data.get('invoice_number')
        remarks = request.data.get('remarks', '')
        invoice_status = request.data.get('status', 'CONFIRMED')
        company_id = request.data.get('company')  # optional — auto-derived from items
        
        if not all([sales_order_id, invoice_date]):
            return Response(
                {'error': 'sales_order and invoice_date are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from Dispatch.models import DispatchOrderItem
            from django.db.models import Sum
            from django.db import transaction
            from Masters.models import Location
            from collections import defaultdict
            from datetime import datetime
            
            sales_order_qs = SalesOrder.filtered_objects.get_qs(
                user=request.user,
                id=sales_order_id
            ).prefetch_related('items', 'items__company')
            sales_order_qs = apply_company_location_filter(
                sales_order_qs, request.user, company_field='items__company'
            )
            sales_order = sales_order_qs.get()

            # Location is optional
            location = None
            if location_id:
                location_qs = apply_company_location_filter(
                    Location.objects.filter(id=location_id, is_deleted=False),
                    request.user,
                    company_field='companies',
                    location_field='id'
                )
                location = location_qs.get()
            
            invoice_date_obj = datetime.strptime(invoice_date, '%Y-%m-%d').date()
            if due_date:
                due_date_obj = datetime.strptime(due_date, '%Y-%m-%d').date()
            else:
                due_date_obj = invoice_date_obj + timedelta(days=sales_order.credit_days)
            
            created_by_type = 'User'
            created_by_id = str(request.user.id) if request.user.is_authenticated else 'Anonymous'
            
            # Collect available items and group by company
            company_items = defaultdict(list)
            for so_item in sales_order.items.all():
                dispatched_qty = DispatchOrderItem.objects.filter(
                    sales_order_item=so_item,
                    dispatch_item__dispatch_plan__status__in=['CONFIRMED', 'DELIVERED']
                ).aggregate(total=Sum('quantity_dispatched'))['total'] or 0
                
                invoiced_qty = InvoiceItem.objects.filter(
                    sales_order_item=so_item,
                    invoice__source_type='DISPATCH',
                    invoice__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID']
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                available_qty = dispatched_qty - invoiced_qty
                if available_qty > 0:
                    item_company_id = so_item.company_id or company_id
                    company_items[item_company_id].append((so_item, available_qty))
            
            if not company_items:
                return Response(
                    {'error': 'No available dispatched quantities to invoice'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_invoices = []
            with transaction.atomic():
                for comp_id, items_list in company_items.items():
                    if not comp_id:
                        return Response(
                            {'error': 'Some order items have no company assigned. Cannot create invoice.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    if len(company_items) == 1 and invoice_number:
                        inv_number = invoice_number
                    elif location:
                        inv_number = Invoice.generate_invoice_number(location.code)
                    else:
                        inv_number = Invoice.generate_invoice_number()
                    
                    invoice = Invoice.objects.create(
                        invoice_number=inv_number,
                        invoice_date=invoice_date_obj,
                        due_date=due_date_obj,
                        source_type='DISPATCH',
                        sales_order=sales_order,
                        company_id=comp_id,
                        location=location,
                        status=invoice_status,
                        remarks=remarks,
                        created_by_type=created_by_type,
                        created_by_identifier=created_by_id
                    )
                    
                    for so_item, available_qty in items_list:
                        ratio = Decimal(str(available_qty)) / Decimal(str(so_item.quantity))
                        InvoiceItem.objects.create(
                            invoice=invoice,
                            sales_order_item=so_item,
                            quantity=available_qty,
                            rate=so_item.rate,
                            discount_amount=so_item.discount_amount * ratio,
                            taxable_amount=so_item.taxable_amount * ratio,
                            cgst_rate=so_item.cgst_rate,
                            cgst_amount=so_item.cgst_amount * ratio,
                            sgst_rate=so_item.sgst_rate,
                            sgst_amount=so_item.sgst_amount * ratio,
                            igst_rate=so_item.igst_rate,
                            igst_amount=so_item.igst_amount * ratio,
                            cess_rate=so_item.cess_rate,
                            cess_amount=so_item.cess_amount * ratio,
                            tax_amount=so_item.tax_amount * ratio,
                            line_total=so_item.line_total * ratio,
                            created_by_type=created_by_type,
                            created_by_identifier=created_by_id
                        )
                    
                    invoice.calculate_totals()
                    if invoice.status == 'DRAFT':
                        _set_draft_authorization(invoice)
                    else:
                        _sync_invoice_status_from_authorization(invoice)
                    created_invoices.append(invoice)
                
                sales_order.update_invoice_status()
            
            serializer = InvoiceDetailSerializer(created_invoices, many=True)
            # Return single object for backward compatibility when only 1 invoice
            if len(created_invoices) == 1:
                return Response(serializer.data[0], status=status.HTTP_201_CREATED)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except SalesOrder.DoesNotExist:
            return Response({'error': 'Sales order not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GenerateInvoiceFromOrderView(APIView):
    """Generate invoice(s) from sales order.
    Items are grouped by their company — one invoice per company."""
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def post(self, request):
        # Handle both FormData and JSON
        data = {}
        for key, value in request.data.items():
            if isinstance(value, list) and len(value) == 1:
                data[key] = value[0]
            else:
                data[key] = value
        
        sales_order_id = data.get('sales_order_id') or data.get('sales_order')
        location_id = data.get('location')
        invoice_date = data.get('invoice_date')
        due_date = data.get('due_date')
        invoice_status = data.get('status', 'CONFIRMED')
        company_id = data.get('company')  # optional — auto-derived from items
        
        if not all([sales_order_id, invoice_date]):
            return Response(
                {'error': 'sales_order and invoice_date are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from django.db.models import Sum
            from django.db import transaction
            from Masters.models import Location
            from collections import defaultdict
            from datetime import datetime

            sales_order_qs = SalesOrder.filtered_objects.get_qs(
                user=request.user,
                id=sales_order_id
            ).prefetch_related('items', 'items__company')
            sales_order_qs = apply_company_location_filter(
                sales_order_qs, request.user, company_field='items__company'
            )
            sales_order = sales_order_qs.get()
            
            # Check if already invoiced from order (including DRAFT status)
            if sales_order.invoices.filter(
                source_type='ORDER',
                status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID']
            ).exists():
                return Response(
                    {'error': 'This sales order is already invoiced from ORDER source'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Location is optional
            location = None
            if location_id:
                location_qs = apply_company_location_filter(
                    Location.objects.filter(id=location_id, is_deleted=False),
                    request.user,
                    company_field='companies',
                    location_field='id'
                )
                location = location_qs.get()
            
            if isinstance(invoice_date, str):
                invoice_date_obj = datetime.strptime(invoice_date, '%Y-%m-%d').date()
            else:
                invoice_date_obj = invoice_date
            
            if not due_date:
                due_date_obj = invoice_date_obj + timedelta(days=sales_order.credit_days)
            else:
                due_date_obj = datetime.strptime(due_date, '%Y-%m-%d').date() if isinstance(due_date, str) else due_date
            
            created_by_type = 'User'
            created_by_id = str(request.user.id) if request.user.is_authenticated else 'Anonymous'
            
            # Collect available items and group by company
            company_items = defaultdict(list)
            for so_item in sales_order.items.all():
                invoiced_qty = InvoiceItem.objects.filter(
                    sales_order_item=so_item,
                    invoice__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID']
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
                
                remaining_qty = so_item.quantity - invoiced_qty
                if remaining_qty > 0:
                    item_company_id = so_item.company_id or company_id
                    company_items[item_company_id].append((so_item, remaining_qty))
            
            if not company_items:
                return Response(
                    {'error': 'No remaining quantities to invoice'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_invoices = []
            is_first = True
            with transaction.atomic():
                for comp_id, items_list in company_items.items():
                    if not comp_id:
                        return Response(
                            {'error': 'Some order items have no company assigned. Cannot create invoice.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    inv_number = Invoice.generate_invoice_number(location.code) if location else Invoice.generate_invoice_number()
                    
                    invoice = Invoice.objects.create(
                        invoice_number=inv_number,
                        invoice_date=invoice_date_obj,
                        due_date=due_date_obj,
                        source_type='ORDER',
                        sales_order=sales_order,
                        company_id=comp_id,
                        location=location,
                        status=invoice_status,
                        freight_charges=sales_order.freight_charges if is_first else 0,
                        other_charges=sales_order.other_charges if is_first else 0,
                        round_off=sales_order.round_off if is_first else 0,
                        created_by_type=created_by_type,
                        created_by_identifier=created_by_id
                    )
                    is_first = False
                    
                    for so_item, remaining_qty in items_list:
                        ratio = remaining_qty / so_item.quantity
                        InvoiceItem.objects.create(
                            invoice=invoice,
                            sales_order_item=so_item,
                            quantity=remaining_qty,
                            rate=so_item.rate,
                            discount_amount=so_item.discount_amount * ratio,
                            taxable_amount=so_item.taxable_amount * ratio,
                            cgst_rate=so_item.cgst_rate,
                            cgst_amount=so_item.cgst_amount * ratio,
                            sgst_rate=so_item.sgst_rate,
                            sgst_amount=so_item.sgst_amount * ratio,
                            igst_rate=so_item.igst_rate,
                            igst_amount=so_item.igst_amount * ratio,
                            cess_rate=so_item.cess_rate,
                            cess_amount=so_item.cess_amount * ratio,
                            tax_amount=so_item.tax_amount * ratio,
                            line_total=so_item.line_total * ratio,
                            created_by_type=created_by_type,
                            created_by_identifier=created_by_id
                        )
                    
                    invoice.calculate_totals()
                    if invoice.status == 'DRAFT':
                        _set_draft_authorization(invoice)
                    else:
                        _sync_invoice_status_from_authorization(invoice)
                    created_invoices.append(invoice)
                
                sales_order.update_invoice_status()
            
            serializer = InvoiceDetailSerializer(created_invoices, many=True)
            if len(created_invoices) == 1:
                return Response(serializer.data[0], status=status.HTTP_201_CREATED)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except SalesOrder.DoesNotExist:
            return Response({'error': 'Sales order not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AvailableDispatchPlansView(generics.ListAPIView):
    """Get dispatch plans available for invoicing"""
    permission_classes = [permissions.AllowAny]
    serializer_class = DispatchPlanForInvoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['location']
    search_fields = ['dispatch_number']
    
    def get_queryset(self):
        accessible_orders = SalesOrder.filtered_objects.get_qs(
            user=self.request.user,
            is_deleted=False
        )
        accessible_orders = apply_company_location_filter(
            accessible_orders,
            self.request.user,
            company_field='company'
        )

        # CONFIRMED or DELIVERED dispatch plans that are not yet invoiced
        queryset = DispatchPlan.objects.filter(
            status__in=['CONFIRMED', 'DELIVERED'],
            is_deleted=False,
            items__sales_order__in=accessible_orders
        ).select_related('company', 'location').prefetch_related(
            'items__sales_order', 'invoices'
        )
        queryset = apply_company_location_filter(
            queryset, self.request.user,
            company_field='company', location_field='location'
        )

        # Filter by customer if provided
        customer_id = self.request.query_params.get('customer')
        customer_type = self.request.query_params.get('customer_type')
        
        if customer_id:
            queryset = queryset.filter(items__sales_order__in=accessible_orders.filter(
                **{f'{customer_type.lower()}': customer_id} if customer_type else {}
            ))
        elif customer_type:
            queryset = queryset.filter(items__sales_order__customer_type=customer_type)
        
        # Exclude already invoiced dispatch plans
        from django.db.models import Q
        queryset = queryset.exclude(
            Q(invoices__status__in=['CONFIRMED', 'PAID', 'PARTIALLY_PAID'])
        ).distinct()
        
        return queryset


class AvailableOrdersForInvoiceView(generics.ListAPIView):
    """Get sales orders available for invoicing"""
    permission_classes = [permissions.AllowAny]
    serializer_class = SalesOrderForInvoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['customer_type']
    search_fields = ['order_number']
    
    def get_queryset(self):
        source_type = self.request.query_params.get('source_type', 'ORDER')
        
        if source_type == 'DISPATCH':
            # For DISPATCH: Only orders with dispatched items not yet fully invoiced
            from django.db.models import Sum, Q, OuterRef, Subquery, DecimalField
            from django.db.models.functions import Coalesce
            from Dispatch.models import DispatchOrderItem
            
            # Filter by customer if provided
            customer_id = self.request.query_params.get('customer')
            customer_type = self.request.query_params.get('customer_type')
            
            # Subquery to get orders with available dispatched quantities
            # Calculate dispatched quantity per order item
            dispatched_subquery = DispatchOrderItem.objects.filter(
                sales_order_item=OuterRef('pk'),
                dispatch_item__dispatch_plan__status__in=['CONFIRMED', 'DELIVERED']
            ).values('sales_order_item').annotate(
                total=Coalesce(Sum('quantity_dispatched'), Decimal('0'))
            ).values('total')[:1]
            
            # Calculate invoiced quantity per order item (including DRAFT status)
            invoiced_subquery = InvoiceItem.objects.filter(
                sales_order_item=OuterRef('pk'),
                invoice__source_type='DISPATCH',
                invoice__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID']
            ).values('sales_order_item').annotate(
                total=Coalesce(Sum('quantity'), Decimal('0'))
            ).values('total')[:1]
            
            # Annotate order items with dispatched and invoiced quantities
            items_with_available = SalesOrderItem.filtered_objects.get_qs(
                user=self.request.user
            ).annotate(
                dispatched_qty=Coalesce(Subquery(dispatched_subquery, output_field=DecimalField()), Decimal('0')),
                invoiced_qty=Coalesce(Subquery(invoiced_subquery, output_field=DecimalField()), Decimal('0'))
            ).filter(
                dispatched_qty__gt=F('invoiced_qty')
            ).values_list('order', flat=True).distinct()
            
            queryset = SalesOrder.filtered_objects.get_qs(
                user=self.request.user,
                id__in=items_with_available,
                authorized_status=2,  # Approved
                is_deleted=False
            ).select_related(
                'retailer', 'distributor', 'superstockist'
            ).prefetch_related('items')
            queryset = apply_company_location_filter(queryset, self.request.user, company_field='company')

            if customer_id and customer_type:
                queryset = queryset.filter(**{f'{customer_type.lower()}': customer_id})
            elif customer_type:
                queryset = queryset.filter(customer_type=customer_type)
            
            return queryset
        
        else:
            # For ORDER: Only authorized orders not yet invoiced from ORDER source
            queryset = SalesOrder.filtered_objects.get_qs(
                user=self.request.user,
                authorized_status=2,  # Approved
                is_deleted=False
            ).select_related(
                'retailer', 'distributor', 'superstockist'
            ).prefetch_related('invoices')
            queryset = apply_company_location_filter(queryset, self.request.user, company_field='company')

            # Filter by customer if provided
            customer_id = self.request.query_params.get('customer')
            customer_type = self.request.query_params.get('customer_type')
            
            if customer_id and customer_type:
                queryset = queryset.filter(**{f'{customer_type.lower()}': customer_id})
            elif customer_type:
                queryset = queryset.filter(customer_type=customer_type)
            
            # Exclude already invoiced orders from ORDER source (with DRAFT or higher status)
            from django.db.models import Q
            queryset = queryset.exclude(
                Q(invoices__source_type='ORDER') & Q(invoices__status__in=['DRAFT', 'CONFIRMED', 'PAID', 'PARTIALLY_PAID'])
            )
            
            return queryset
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        
        source_type = request.query_params.get('source_type', 'ORDER')
        
        if page is not None:
            serializer = self.get_serializer(page, many=True, context={'source_type': source_type})
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True, context={'source_type': source_type})
        return Response(serializer.data)


class GenerateInvoiceNumberView(APIView):
    """Generate new invoice number"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        location_id = request.query_params.get('location')
        if not location_id:
            return Response(
                {'error': 'Location parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from Masters.models import Location
            location_qs = apply_company_location_filter(
                Location.objects.filter(id=location_id, is_deleted=False),
                request.user,
                company_field='companies',
                location_field='id'
            )
            location = location_qs.get()
            invoice_number = Invoice.generate_invoice_number(location.code)
            return Response({'invoice_number': invoice_number})
        except Location.DoesNotExist:
            return Response({'error': 'Invalid location'}, status=status.HTTP_400_BAD_REQUEST)


class CancelInvoiceView(APIView):
    """Cancel an invoice"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        try:
            qs = Invoice.filtered_objects.get_qs(user=request.user)
            qs = apply_company_location_filter(qs, request.user, company_field='company', location_field='location')
            invoice = qs.get(id=pk)

            # Only allow cancellation of DRAFT or CONFIRMED invoices
            if invoice.status not in ['DRAFT', 'CONFIRMED']:
                return Response(
                    {'error': 'Only DRAFT or CONFIRMED invoices can be cancelled'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if invoice has payments
            if invoice.paid_amount > 0:
                return Response(
                    {'error': 'Cannot cancel invoice with payments. Please reverse payments first.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Cancel the invoice
            invoice.status = 'CANCELLED'
            invoice.save()
            
            # Update sales order status
            invoice.sales_order.update_invoice_status()
            
            return Response(
                {'message': 'Invoice cancelled successfully'},
                status=status.HTTP_200_OK
            )
            
        except Invoice.DoesNotExist:
            return Response({'error': 'Invoice not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GetCustomerPendingInvoicesView(APIView):
    """Get pending invoices for a customer with balance > 0"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        """
        Fetch pending invoices for a customer
        Query params: customer_type, customer_id
        Returns invoices where balance_amount > 0 and status != CANCELLED
        """
        customer_type = request.query_params.get('customer_type')
        customer_id = request.query_params.get('customer_id')
        
        if not customer_type or not customer_id:
            return Response(
                {'error': 'Both customer_type and customer_id parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build filter based on customer type
        filter_kwargs = {
            'status__in': ['CONFIRMED', 'PAID', 'PARTIALLY_PAID'],  # Exclude DRAFT and CANCELLED
            'balance_amount__gt': 0  # Only unpaid/partially paid invoices
        }
        
        # Map customer type to filter field
        if customer_type == 'RETAILER':
            filter_kwargs['sales_order__retailer_id'] = customer_id
        elif customer_type == 'DISTRIBUTOR':
            filter_kwargs['sales_order__distributor_id'] = customer_id
        elif customer_type == 'SUPERSTOCKIST':
            filter_kwargs['sales_order__superstockist_id'] = customer_id
        else:
            return Response(
                {'error': f'Invalid customer_type: {customer_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Fetch pending invoices filtered by user's company/location access
            invoices = Invoice.filtered_objects.get_qs(
                user=request.user,
                **filter_kwargs
            ).select_related(
                'sales_order'
            ).order_by('-invoice_date')
            invoices = apply_company_location_filter(
                invoices, request.user,
                company_field='company', location_field='location'
            )
            
            # Serialize data with calculated days
            data = []
            today = timezone.now().date()
            
            for invoice in invoices:
                days_diff = (today - invoice.invoice_date).days
                data.append({
                    'id': str(invoice.id),
                    'invoice_number': invoice.invoice_number,
                    'invoice_date': invoice.invoice_date.strftime('%d-%m-%Y'),  # Format as DD-MM-YYYY
                    'grand_total': float(invoice.grand_total),
                    'balance_amount': float(invoice.balance_amount),
                    'days': days_diff
                })
            
            return Response({
                'count': len(data),
                'results': data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching pending invoices: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )




class InvoiceStatusCountView(generics.ListAPIView):
    """Get count of invoices grouped by status"""
    permission_classes = [permissions.AllowAny]
    serializer_class = InvoiceStatusCountSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = InvoiceFilter

    def get_queryset(self):
        queryset = Invoice.objects.filter(is_deleted=False)
        queryset = apply_company_location_filter(
            queryset,
            self.request.user,
            company_field='company',
            location_field='location'
        )
        return queryset

    def list(self, request, *args, **kwargs):
        try:
            # Get filtered queryset (applies all filter backends and filterset)
            queryset = self.filter_queryset(self.get_queryset())
            
            # Get counts by status
            status_counts = queryset.values('status').annotate(
                count=Count('id')
            ).order_by('status')
            
            # Build response with all statuses
            all_statuses = {
                'DRAFT': 0,
                'PENDING': 0,
                'CONFIRMED': 0,
                'PAID': 0,
                'PARTIALLY_PAID': 0,
                'CANCELLED': 0,
            }
            
            # Populate with actual counts
            for item in status_counts:
                if item['status'] in all_statuses:
                    all_statuses[item['status']] = item['count']
            
            # Add total count
            all_statuses['total'] = sum(all_statuses.values())
            
            # Serialize the response
            serializer = self.get_serializer(all_statuses)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error fetching status counts: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        


# ============================================================
# Invoice Report Views
# ============================================================

def _get_invoice_report_queryset(user):
    """Shared queryset builder for invoice report view and export view."""
    queryset = Invoice.filtered_objects.get_qs(
        user=user
    ).select_related(
        'sales_order',
        'sales_order__billing_state',
        'sales_order__billing_state__country',
        'sales_order__billing_city',
        'sales_order__billing_city__district',
        'sales_order__billing_city__mandal',
        'sales_order__billing_area',
        'sales_order__retailer',
        'sales_order__distributor',
        'sales_order__distributor__agent',
        'sales_order__superstockist',
        'dispatch_plan',
        'company',
        'location'
    )
    return apply_company_location_filter(queryset, user, company_field='company', location_field='location')


def _convert_to_invoice_rows(invoices):
    """Convert invoices to report rows with all required fields."""
    report_rows = []

    for invoice in invoices:
        so = invoice.sales_order
        billing_state = so.billing_state
        billing_city = so.billing_city
        billing_area = so.billing_area

        source_type_display = dict(invoice._meta.get_field('source_type').choices).get(invoice.source_type, invoice.source_type)
        status_display = dict(invoice._meta.get_field('status').choices).get(invoice.status, invoice.status)
        pod_status_display = dict(invoice._meta.get_field('pod_status').choices).get(invoice.pod_status, invoice.pod_status)
        
        # Authorization status mapping
        authorization_status_choices = {
            1: 'PENDING',
            2: 'APPROVED', 
            3: 'REJECTED'
        }
        authorization_status_code = authorization_status_choices.get(invoice.authorized_status, 'PENDING') if invoice.authorized_status else 'PENDING'

        report_rows.append({
            'id': str(invoice.id),
            'invoice_number': invoice.invoice_number,
            'invoice_date': invoice.invoice_date,
            'due_date': invoice.due_date,
            'source_type': invoice.source_type,
            'source_type_display': source_type_display,
            'dispatch_number': invoice.dispatch_plan.dispatch_number if invoice.dispatch_plan else None,
            'sales_order_number': so.order_number,
            'customer_name': so.get_customer_name(),
            'customer_type': so.get_customer_type_display(),
            'customer_type_code': so.customer_type,
            'subtotal': invoice.subtotal,
            'discount_amount': invoice.discount_amount,
            'taxable_amount': invoice.taxable_amount,
            'tax_amount': invoice.tax_amount,
            'freight_charges': invoice.freight_charges,
            'other_charges': invoice.other_charges,
            'round_off': invoice.round_off,
            'grand_total': invoice.grand_total,
            'paid_amount': invoice.paid_amount,
            'balance_amount': invoice.balance_amount,
            'status': status_display,
            'status_code': invoice.status,
            'pod_status': pod_status_display,
            'pod_status_code': invoice.pod_status,
            'authorization_status': invoice.authorized_status,
            'authorization_status_code': authorization_status_code,
            'authorization_level': invoice.authorized_level,
            'current_authorized_level': invoice.current_authorized_level,
            'country': billing_state.country.name if billing_state and billing_state.country else None,
            'state': billing_state.name if billing_state else '',
            'state_id': str(billing_state.id) if billing_state else None,
            'district': billing_city.district.name if billing_city and billing_city.district else None,
            'mandal': billing_city.mandal.name if billing_city and billing_city.mandal else None,
            'city': billing_city.name if billing_city else '',
            'city_id': str(billing_city.id) if billing_city else None,
            'area': billing_area.name if billing_area else None,
            'area_id': str(billing_area.id) if billing_area else None,
            'company_name': invoice.company.name,
            'location_name': invoice.location.name if invoice.location else None,
            'agent_name': so.distributor.agent.name if so.distributor and so.distributor.agent else None,
            'created_on': invoice.created_on,
            'modified_on': invoice.modified_on,
        })

    return report_rows


_INVOICE_EXPORT_HEADERS = [
    'Invoice Number', 'Invoice Date', 'Due Date',
    'Source Type', 'Dispatch Number', 'Sales Order Number',
    'Customer Name', 'Customer Type',
    'Subtotal', 'Discount', 'Taxable Amount', 'Tax Amount',
    'Freight', 'Other Charges', 'Round Off', 'Grand Total',
    'Paid Amount', 'Balance Amount',
    'Status', 'POD Status',
    'State', 'City', 'Area',
    'Company', 'Location',
]


def _iter_invoice_export_rows(queryset):
    """Yield one flat row tuple per invoice."""
    for invoice in queryset:
        so = invoice.sales_order
        yield (
            invoice.invoice_number,
            invoice.invoice_date.strftime('%Y-%m-%d'),
            invoice.due_date.strftime('%Y-%m-%d'),
            dict(invoice._meta.get_field('source_type').choices).get(invoice.source_type, invoice.source_type),
            invoice.dispatch_plan.dispatch_number if invoice.dispatch_plan else '',
            so.order_number,
            so.get_customer_name(),
            so.get_customer_type_display(),
            float(invoice.subtotal),
            float(invoice.discount_amount),
            float(invoice.taxable_amount),
            float(invoice.tax_amount),
            float(invoice.freight_charges),
            float(invoice.other_charges),
            float(invoice.round_off),
            float(invoice.grand_total),
            float(invoice.paid_amount),
            float(invoice.balance_amount),
            dict(invoice._meta.get_field('status').choices).get(invoice.status, invoice.status),
            dict(invoice._meta.get_field('pod_status').choices).get(invoice.pod_status, invoice.pod_status),
            so.billing_state.name if so.billing_state else '',
            so.billing_city.name if so.billing_city else '',
            so.billing_area.name if so.billing_area else '',
            invoice.company.name,
            invoice.location.name if invoice.location else '',
        )


class InvoiceReportFilter(FilterSet):
    """Advanced filter for Invoice Report."""
    from_date = DateFilter(field_name='invoice_date', lookup_expr='gte', label='From Date')
    to_date = DateFilter(field_name='invoice_date', lookup_expr='lte', label='To Date')
    date_preset = CharFilter(method='filter_date_preset', label='Date Preset')
    due_from_date = DateFilter(field_name='due_date', lookup_expr='gte', label='Due From Date')
    due_to_date = DateFilter(field_name='due_date', lookup_expr='lte', label='Due To Date')
    source_type = ChoiceFilter(
        field_name='source_type',
        choices=[('DISPATCH', 'From Dispatch Plan'), ('ORDER', 'From Sales Order')],
        label='Source Type'
    )
    dispatch_number = CharFilter(method='filter_dispatch_number', label='Dispatch Number')
    sales_order_number = CharFilter(method='filter_sales_order_number', label='Sales Order Number')
    customer_type = ChoiceFilter(
        method='filter_customer_type',
        choices=[('RETAILER', 'Retailer'), ('DISTRIBUTOR', 'Distributor'), ('SUPERSTOCKIST', 'Superstockist')],
        label='Customer Type'
    )
    customer_id = CharFilter(method='filter_customer', label='Customer ID')
    country = UUIDFilter(method='filter_country', label='Country')
    state = UUIDFilter(method='filter_state', label='State')
    district = UUIDFilter(method='filter_district', label='District')
    mandal = UUIDFilter(method='filter_mandal', label='Mandal')
    city = UUIDFilter(method='filter_city', label='City')
    area = UUIDFilter(method='filter_area', label='Area')
    invoice_status = ChoiceFilter(
        field_name='status',
        choices=[
            ('DRAFT', 'Draft'), ('PENDING', 'Pending'), ('CONFIRMED', 'Confirmed'),
            ('PAID', 'Paid'), ('PARTIALLY_PAID', 'Partially Paid'), ('CANCELLED', 'Cancelled'),
        ],
        label='Invoice Status'
    )
    authorization_status = NumberFilter(
        field_name='authorized_status',
        label='Authorization Status'
    )
    pod_status = ChoiceFilter(
        field_name='pod_status',
        choices=[('PENDING', 'POD Pending'), ('COMPLETED', 'POD Completed')],
        label='POD Status'
    )
    agent = UUIDFilter(field_name='sales_order__distributor__agent_id', label='Agent')

    class Meta:
        model = Invoice
        fields = [
            'from_date', 'to_date', 'date_preset',
            'due_from_date', 'due_to_date',
            'source_type', 'dispatch_number', 'sales_order_number',
            'customer_type', 'customer_id',
            'country', 'state', 'district', 'mandal', 'city', 'area',
            'invoice_status', 'authorization_status', 'pod_status', 'agent'
        ]

    def filter_date_preset(self, queryset, name, value):
        today = date.today()
        presets = {
            'today': (today, today),
            'this_week': (today - timedelta(days=today.weekday()), today),
            'this_month': (today.replace(day=1), today),
            'this_year': (today.replace(month=1, day=1), today),
        }
        date_range = presets.get(value)
        if date_range:
            return queryset.filter(invoice_date__gte=date_range[0], invoice_date__lte=date_range[1])
        return queryset

    def filter_dispatch_number(self, queryset, name, value):
        return queryset.filter(dispatch_plan__dispatch_number__icontains=value)

    def filter_sales_order_number(self, queryset, name, value):
        return queryset.filter(sales_order__order_number__icontains=value)

    def filter_customer_type(self, queryset, name, value):
        return queryset.filter(sales_order__customer_type=value)

    def filter_customer(self, queryset, name, value):
        return queryset.filter(
            Q(sales_order__retailer_id=value) |
            Q(sales_order__distributor_id=value) |
            Q(sales_order__superstockist_id=value)
        )

    def filter_country(self, queryset, name, value):
        return queryset.filter(sales_order__billing_state__country_id=value)

    def filter_state(self, queryset, name, value):
        return queryset.filter(sales_order__billing_state_id=value)

    def filter_district(self, queryset, name, value):
        return queryset.filter(sales_order__billing_city__district_id=value)

    def filter_mandal(self, queryset, name, value):
        return queryset.filter(sales_order__billing_city__mandal_id=value)

    def filter_city(self, queryset, name, value):
        return queryset.filter(sales_order__billing_city_id=value)

    def filter_area(self, queryset, name, value):
        return queryset.filter(sales_order__billing_area_id=value)

    def filter_authorization_status(self, queryset, name, value):
        """Filter by authorization status"""
        try:
            status_int = int(value)
            return queryset.filter(authorized_status=status_int)
        except (ValueError, TypeError):
            return queryset


class InvoiceReportView(generics.ListAPIView):
    """
    Invoice Report API.

    GET /api/invoice/reports/invoices/
    """

    permission_classes = [permissions.AllowAny]
    serializer_class = InvoiceReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = InvoiceReportFilter

    search_fields = [
        'invoice_number', 'code',
        'sales_order__order_number',
        'dispatch_plan__dispatch_number',
        'sales_order__retailer__name', 'sales_order__retailer__code',
        'sales_order__distributor__name', 'sales_order__distributor__code',
        'sales_order__superstockist__name', 'sales_order__superstockist__code',
        'sales_order__billing_city__name', 'sales_order__billing_state__name',
        'sales_order__billing_area__name',
    ]

    ordering_fields = [
        'invoice_date', 'due_date', 'invoice_number',
        'grand_total', 'status', 'balance_amount',
    ]

    ordering = ['-invoice_date', 'invoice_number']

    def get_queryset(self):
        return _get_invoice_report_queryset(self.request.user)

    def _get_summary(self, queryset):
        """Compute summary stats using DB aggregation."""
        agg = queryset.aggregate(
            total_invoices=Count('id'),
            total_amount=Sum('grand_total'),
            total_tax=Sum('tax_amount'),
            total_paid=Sum('paid_amount'),
            total_balance=Sum('balance_amount'),
        )
        return {
            'total_invoices': agg['total_invoices'] or 0,
            'total_amount': float(agg['total_amount'] or 0),
            'total_tax': float(agg['total_tax'] or 0),
            'total_paid': float(agg['total_paid'] or 0),
            'total_balance': float(agg['total_balance'] or 0),
        }

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        summary = self._get_summary(queryset)

        page = self.paginate_queryset(queryset)
        if page is not None:
            report_rows = _convert_to_invoice_rows(page)
            serializer = self.get_serializer(report_rows, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['summary'] = summary
            return response

        report_rows = _convert_to_invoice_rows(queryset)
        serializer = self.get_serializer(report_rows, many=True)
        return Response({
            'results': serializer.data,
            'summary': summary,
        })


class InvoiceReportExportView(APIView):
    """
    Export Invoice Report to Excel, CSV, or PDF.

    POST /api/invoice/reports/invoices/export/
    """

    permission_classes = [permissions.AllowAny]

    VALID_FORMATS = frozenset(['excel', 'csv', 'pdf'])

    def post(self, request):
        export_format = request.data.get('format', 'excel').lower()
        report_filters = request.data.get('filters', {})

        if export_format not in self.VALID_FORMATS:
            return Response(
                {'error': 'Invalid format. Must be excel, csv, or pdf'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = self._get_filtered_queryset(request, report_filters)
        handler = getattr(self, f'_export_{export_format}')
        return handler(queryset)

    def _get_filtered_queryset(self, request, report_filters):
        queryset = _get_invoice_report_queryset(request.user)
        filterset = InvoiceReportFilter(report_filters, queryset=queryset, request=request)
        return filterset.qs

    def _export_excel(self, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = 'Invoice Report'

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(_INVOICE_EXPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, row_data in enumerate(_iter_invoice_export_rows(queryset), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=invoice_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    def _export_csv(self, queryset):
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=invoice_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        writer = csv.writer(response)
        writer.writerow(_INVOICE_EXPORT_HEADERS)

        for row in _iter_invoice_export_rows(queryset):
            writer.writerow(row)

        return response

    def _export_pdf(self, queryset):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1,
        )

        elements.append(Paragraph('Invoice Report', title_style))
        elements.append(Spacer(1, 0.2 * inch))

        pdf_headers = [
            'Invoice #', 'Date', 'Customer', 'Type',
            'Grand Total', 'Paid', 'Balance', 'Status',
        ]
        data = [pdf_headers]

        max_pdf_rows = 100
        total_rows = queryset.count()
        if total_rows > max_pdf_rows:
            return Response(
                {'error': f'PDF export is limited to {max_pdf_rows} invoices. '
                          f'Your filter returned {total_rows} invoices. '
                          f'Please narrow your filters or use CSV/Excel export for larger datasets.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for full_row in _iter_invoice_export_rows(queryset):
            data.append([
                full_row[0],                        # Invoice Number
                full_row[1],                        # Invoice Date
                full_row[6][:20],                   # Customer Name (truncated)
                full_row[7][:15],                   # Customer Type (truncated)
                f"\u20B9{full_row[15]:,.2f}",       # Grand Total
                f"\u20B9{full_row[16]:,.2f}",       # Paid Amount
                f"\u20B9{full_row[17]:,.2f}",       # Balance Amount
                full_row[18][:12],                  # Status (truncated)
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.write(pdf)
        return response
