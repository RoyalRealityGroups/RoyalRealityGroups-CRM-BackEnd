"""
Module 11 - Report Services
Generates Lead, Site Visit, and Sales reports with Excel/PDF export.
"""
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import timedelta, date


def _date_range_filter(period: str, date_field: str = 'created_on') -> dict:
    """Return Django filter kwargs for a named period."""
    today = timezone.now().date()
    if period == 'today':
        return {f'{date_field}__date': today}
    if period == 'this_week':
        start = today - timedelta(days=today.weekday())
        return {f'{date_field}__date__gte': start}
    if period == 'this_month':
        return {f'{date_field}__year': today.year, f'{date_field}__month': today.month}
    if period == 'last_month':
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return {f'{date_field}__year': last_month.year, f'{date_field}__month': last_month.month}
    if period == 'this_year':
        return {f'{date_field}__year': today.year}
    return {}


# ============================================================================
# LEAD REPORTS
# ============================================================================

def lead_report_by_source(period=None, project_id=None):
    """Lead count grouped by source."""
    from Lead.models import Lead
    qs = Lead.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period))
    if project_id:
        qs = qs.filter(interested_project_id=project_id)
    data = list(
        qs.values('lead_source')
          .annotate(count=Count('id'))
          .order_by('-count')
    )
    total = qs.count()
    return {'total': total, 'data': data}


def lead_report_by_employee(period=None, project_id=None):
    """Lead count grouped by assigned employee."""
    from Lead.models import Lead
    qs = Lead.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period))
    if project_id:
        qs = qs.filter(interested_project_id=project_id)
    data = list(
        qs.values(
            'assigned_employee__id',
            'assigned_employee__first_name',
            'assigned_employee__last_name',
            'assigned_employee__username',
        ).annotate(count=Count('id')).order_by('-count')
    )
    # normalise employee name
    for row in data:
        fn = row.pop('assigned_employee__first_name', '') or ''
        ln = row.pop('assigned_employee__last_name', '') or ''
        un = row.pop('assigned_employee__username', '') or ''
        eid = row.pop('assigned_employee__id', None)
        row['employee_id'] = str(eid) if eid else None
        row['employee_name'] = f"{fn} {ln}".strip() or un
    return {'total': sum(r['count'] for r in data), 'data': data}


def lead_report_by_project(period=None):
    """Lead count grouped by interested project."""
    from Lead.models import Lead
    qs = Lead.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period))
    data = list(
        qs.values('interested_project__id', 'interested_project__name')
          .annotate(count=Count('id'))
          .order_by('-count')
    )
    for row in data:
        row['project_id'] = str(row.pop('interested_project__id', None) or '')
        row['project_name'] = row.pop('interested_project__name', None) or 'Unassigned'
    return {'total': sum(r['count'] for r in data), 'data': data}


def lead_report_by_status(period=None):
    """Lead count grouped by status."""
    from Lead.models import Lead
    qs = Lead.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period))
    data = list(qs.values('status').annotate(count=Count('id')).order_by('-count'))
    return {'total': qs.count(), 'data': data}


# ============================================================================
# SITE VISIT REPORTS
# ============================================================================

def site_visit_report(period='this_month', project_id=None, employee_id=None):
    """Site visit summary."""
    from SiteVisit.models import SiteVisit
    qs = SiteVisit.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period, date_field='visit_date'))
    if project_id:
        qs = qs.filter(project_id=project_id)
    if employee_id:
        qs = qs.filter(assigned_employee_id=employee_id)

    by_status = {
        item['status']: item['count']
        for item in qs.values('status').annotate(count=Count('id'))
    }
    by_employee = list(
        qs.values(
            'assigned_employee__id',
            'assigned_employee__first_name',
            'assigned_employee__last_name',
        ).annotate(count=Count('id')).order_by('-count')
    )
    for row in by_employee:
        fn = row.pop('assigned_employee__first_name', '') or ''
        ln = row.pop('assigned_employee__last_name', '') or ''
        eid = row.pop('assigned_employee__id', None)
        row['employee_id'] = str(eid) if eid else None
        row['employee_name'] = f"{fn} {ln}".strip() or 'Unassigned'

    return {
        'total': qs.count(),
        'scheduled': by_status.get('SCHEDULED', 0),
        'confirmed': by_status.get('CONFIRMED', 0),
        'completed': by_status.get('COMPLETED', 0),
        'cancelled': by_status.get('CANCELLED', 0),
        'by_employee': by_employee,
    }


# ============================================================================
# SALES / BOOKING REPORTS
# ============================================================================

def booking_report(period=None, project_id=None, employee_id=None):
    """Booking summary with revenue."""
    from Booking.models import Booking
    qs = Booking.objects.filter(is_deleted=False)
    if period:
        qs = qs.filter(**_date_range_filter(period, date_field='booking_date'))
    if project_id:
        qs = qs.filter(project_id=project_id)
    if employee_id:
        qs = qs.filter(sales_executive_id=employee_id)

    by_status = {
        item['status']: item['count']
        for item in qs.values('status').annotate(count=Count('id'))
    }
    active = qs.exclude(status='CANCELLED')
    by_project = list(
        active.values('project__id', 'project__name')
              .annotate(count=Count('id'), revenue=Sum('agreed_price'))
              .order_by('-count')
    )
    for row in by_project:
        row['project_id'] = str(row.pop('project__id', '') or '')
        row['project_name'] = row.pop('project__name', '') or ''
        row['revenue'] = float(row['revenue'] or 0)

    return {
        'total': qs.count(),
        'booked': by_status.get('BOOKED', 0),
        'agreement': by_status.get('AGREEMENT', 0),
        'registered': by_status.get('REGISTERED', 0),
        'cancelled': by_status.get('CANCELLED', 0),
        'total_revenue': float(active.aggregate(t=Sum('agreed_price'))['t'] or 0),
        'booking_collections': float(active.aggregate(t=Sum('booking_amount'))['t'] or 0),
        'by_project': by_project,
    }


def revenue_report(period='this_year'):
    """Monthly revenue breakdown."""
    from Booking.models import Booking
    from django.db.models.functions import TruncMonth
    qs = Booking.objects.filter(is_deleted=False).exclude(status='CANCELLED')
    if period:
        qs = qs.filter(**_date_range_filter(period, date_field='booking_date'))

    monthly = list(
        qs.annotate(month=TruncMonth('booking_date'))
          .values('month')
          .annotate(bookings=Count('id'), revenue=Sum('agreed_price'))
          .order_by('month')
    )
    for row in monthly:
        row['month'] = row['month'].strftime('%Y-%m') if row['month'] else None
        row['revenue'] = float(row['revenue'] or 0)

    return {
        'total_revenue': float(qs.aggregate(t=Sum('agreed_price'))['t'] or 0),
        'monthly': monthly,
    }


def employee_performance_report(period='this_month'):
    """Per-employee: leads, site visits, bookings, registrations."""
    from Users.models import User
    from Lead.models import Lead
    from SiteVisit.models import SiteVisit
    from Booking.models import Booking

    date_filter = _date_range_filter(period)
    sv_filter = _date_range_filter(period, date_field='visit_date')
    bkg_filter = _date_range_filter(period, date_field='booking_date')

    employees = User.objects.filter(is_active=True)
    result = []
    for emp in employees:
        leads = Lead.objects.filter(assigned_employee=emp, is_deleted=False, **date_filter).count()
        visits = SiteVisit.objects.filter(assigned_employee=emp, is_deleted=False, **sv_filter).count()
        bookings = Booking.objects.filter(
            sales_executive=emp, is_deleted=False, **bkg_filter
        ).exclude(status='CANCELLED').count()
        registrations = Booking.objects.filter(
            sales_executive=emp, is_deleted=False, status='REGISTERED', **bkg_filter
        ).count()
        result.append({
            'employee_id': str(emp.id),
            'employee_name': f"{emp.first_name} {emp.last_name}".strip() or emp.username,
            'designation': getattr(emp, 'designation', None),
            'leads': leads,
            'site_visits': visits,
            'bookings': bookings,
            'registrations': registrations,
        })
    result.sort(key=lambda x: x['bookings'], reverse=True)
    return result


# ============================================================================
# EXCEL EXPORT HELPERS
# ============================================================================

def export_to_excel(data: list, columns: list, sheet_name: str = 'Report') -> bytes:
    """Convert list-of-dicts to an Excel file bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['label'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col['key'], ''))

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")
