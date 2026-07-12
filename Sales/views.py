from django.db import models
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count, Sum
from rest_framework import generics, permissions, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters import FilterSet, CharFilter, DateFilter, NumberFilter, UUIDFilter, ChoiceFilter
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, date, timedelta
from decimal import Decimal
import json

from Core.Core.permissions.permissions import GetPermission

from .models import SalesOrder, SalesOrderItem, SalesOrderHistory
from .serializers import (
    SalesOrderSerializer,
    SalesOrderListSerializer,
    SalesOrderItemSerializer,
    SalesOrderHistorySerializer,
    SalesOrderStatusCountSerializer,
    SalesOrderReportSerializer,
)
from .attachment_views import SalesOrderAttachmentMixin
from Masters.models import Scheme
from Masters.serializers import SchemeMiniSerializer
from General.models import GeneralSettings
from utils import apply_company_location_filter


def _set_draft_authorization(instance):
    """Keep draft records out of authorization workflow."""
    instance.authorized_status = None
    instance.current_authorized_status = None
    instance.authorized_level = 0
    instance.current_authorized_level = 0
    instance.authorized_by_type = None
    instance.authorized_by_identifier = None
    instance.authorized_on = None
    instance.current_authorized_by_type = None
    instance.current_authorized_by_identifier = None
    instance.current_authorized_on = None
    instance.save(update_fields=[
        'authorized_status', 'current_authorized_status',
        'authorized_level', 'current_authorized_level',
        'authorized_by_type', 'authorized_by_identifier', 'authorized_on',
        'current_authorized_by_type', 'current_authorized_by_identifier', 'current_authorized_on',
    ])


def _sync_sales_order_status_from_authorization(instance):
    """
    Non-draft orders should reflect authorization progress:
    - authorized_status=APPROVED(2) -> CONFIRMED
    - otherwise -> PENDING
    """
    if instance.status == 'DRAFT':
        return

    target_status = 'CONFIRMED' if instance.authorized_status == 2 else 'PENDING'
    if instance.status != target_status:
        SalesOrder.objects.filter(pk=instance.pk).update(status=target_status)
        instance.status = target_status


class SalesOrderFilter(FilterSet):
    """Filter for Sales Orders"""
    order_number = CharFilter(field_name='order_number', lookup_expr='icontains')
    customer_type = CharFilter(field_name='customer_type')
    status = CharFilter(field_name='status')
    company = CharFilter(field_name='company_id')
    retailer = CharFilter(field_name='retailer_id')
    distributor = CharFilter(field_name='distributor_id')
    superstockist = CharFilter(field_name='superstockist_id')
    order_date_from = DateFilter(field_name='order_date', lookup_expr='gte')
    order_date_to = DateFilter(field_name='order_date', lookup_expr='lte')
    authorized_status = NumberFilter(field_name='authorized_status')
    min_amount = CharFilter(method='filter_min_amount')
    max_amount = CharFilter(method='filter_max_amount')
    
    class Meta:
        model = SalesOrder
        fields = ['order_number', 'customer_type', 'status', 'company', 'retailer', 'distributor', 'superstockist', 'authorized_status', 'order_date_from', 'order_date_to']
    
    def filter_min_amount(self, queryset, name, value):
        return queryset.filter(grand_total__gte=value)
    
    def filter_max_amount(self, queryset, name, value):
        return queryset.filter(grand_total__lte=value)


class SalesOrderList(SalesOrderAttachmentMixin, generics.ListCreateAPIView):
    """List and create sales orders"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = SalesOrderFilter
    search_fields = [
        'order_number',
        'code',
        'remarks',
        'retailer__name', 
        'retailer__code',
        'distributor__name', 
        'distributor__code',
        'superstockist__name',
        'superstockist__code',
        'billing_city__name',
        'billing_state__name',
        'shipping_city__name',
        'shipping_state__name'
    ]
    ordering_fields = ['order_date', 'order_number', 'grand_total', 'authorized_status', 'authorized_on', 'created_on']
    ordering = ['-order_date', '-created_on']
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return SalesOrderSerializer
        return SalesOrderListSerializer
    
    def get_queryset(self):
        queryset = SalesOrder.filtered_objects.get_qs(
            user=self.request.user
        ).select_related(
            'billing_state', 'billing_city', 'billing_area',
            'shipping_state', 'shipping_city', 'shipping_area',
            'retailer', 'distributor', 'superstockist'
        ).prefetch_related('items')

        queryset = apply_company_location_filter(queryset, self.request.user, company_field='company')
        return queryset

    def create(self, request, *args, **kwargs):
        # Convert QueryDict to regular dict and flatten single-item lists
        data = {}
        for key, value in request.data.items():
            if key == 'items' and isinstance(value, str):
                data[key] = json.loads(value)
            elif key == 'selected_scheme_ids' and isinstance(value, str):
                data[key] = json.loads(value)
            elif isinstance(value, list) and len(value) == 1:
                data[key] = value[0]
            else:
                data[key] = value
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_create(self, serializer):
        requested_status = str(self.request.data.get('status', '')).upper()
        instance = serializer.save(
            created_by_type='User',
            created_by_identifier=str(self.request.user.id) if self.request.user.is_authenticated else 'Anonymous'
        )
        if requested_status == 'DRAFT' or (instance.status == 'DRAFT' and requested_status == ''):
            _set_draft_authorization(instance)
        else:
            _sync_sales_order_status_from_authorization(instance)
        # Handle attachments after creating the instance
        self.handle_attachments(instance, self.request)


class SalesOrderDetail(SalesOrderAttachmentMixin, generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a sales order"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    serializer_class = SalesOrderSerializer
    
    def get_queryset(self):
        queryset = SalesOrder.filtered_objects.get_qs(
            user=self.request.user
        ).select_related(
            'billing_state', 'billing_city', 'billing_area',
            'shipping_state', 'shipping_city', 'shipping_area',
            'retailer', 'distributor', 'superstockist'
        ).prefetch_related(
            'items',
            'items__item',
            'items__item__base_uom',
            'items__applied_schemes',
            'applied_schemes',
            'applied_schemes__scheme'
        )
        return apply_company_location_filter(queryset, self.request.user, company_field='company')

    def update(self, request, *args, **kwargs):
        # Check if order status allows editing
        instance = self.get_object()
        if instance.status not in ['DRAFT', 'PENDING', 'CONFIRMED']:
            return Response(
                {'error': f'Cannot edit sales order with status {instance.status}. Only DRAFT, PENDING or CONFIRMED orders can be edited.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Convert QueryDict to regular dict and flatten single-item lists
        data = {}
        for key, value in request.data.items():
            if key == 'items' and isinstance(value, str):
                data[key] = json.loads(value)
            elif key == 'selected_scheme_ids' and isinstance(value, str):
                data[key] = json.loads(value)
            elif isinstance(value, list) and len(value) == 1:
                data[key] = value[0]
            else:
                data[key] = value
        
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    
    def perform_update(self, serializer):
        requested_status = str(self.request.data.get('status', '')).upper()
        instance = serializer.save(
            modified_by_type='User',
            modified_by_identifier=str(self.request.user.id) if self.request.user.is_authenticated else 'Anonymous'
        )
        if requested_status == 'DRAFT':
            if instance.status != 'DRAFT':
                instance.status = 'DRAFT'
                instance.save(update_fields=['status'])
            _set_draft_authorization(instance)
        else:
            _sync_sales_order_status_from_authorization(instance)
        # Handle attachments after updating the instance
        self.handle_attachments(instance, self.request)
    
    def destroy(self, request, *args, **kwargs):
        """Only allow deletion of PENDING or REJECTED orders"""
        instance = self.get_object()
        if instance.status not in ['PENDING', 'REJECTED']:
            return Response(
                {'error': 'Only PENDING or REJECTED orders can be deleted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)


class SalesOrderApprove(APIView):
    """Approve a sales order"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        qs = SalesOrder.filtered_objects.get_qs(user=request.user)
        qs = apply_company_location_filter(qs, request.user, company_field='company')
        order = get_object_or_404(qs, pk=pk)

        if order.status != 'PENDING':
            return Response(
                {'error': 'Only PENDING orders can be approved'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = 'APPROVED'
        order.approved_by = request.user if request.user.is_authenticated else None
        order.approved_at = datetime.now()
        order.save()
        
        # Create history entry
        SalesOrderHistory.objects.create(
            order=order,
            action='APPROVED',
            old_status=old_status,
            new_status=order.status,
            changed_by=request.user if request.user.is_authenticated else None,
            remarks='Order approved'
        )
        
        serializer = SalesOrderSerializer(order)
        return Response(serializer.data)


class SalesOrderReject(APIView):
    """Reject a sales order"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        qs = SalesOrder.filtered_objects.get_qs(user=request.user)
        qs = apply_company_location_filter(qs, request.user, company_field='company')
        order = get_object_or_404(qs, pk=pk)

        if order.status != 'PENDING':
            return Response(
                {'error': 'Only PENDING orders can be rejected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reason = request.data.get('reason', '')
        if not reason:
            return Response(
                {'error': 'Rejection reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = 'REJECTED'
        order.rejection_reason = reason
        order.save()
        
        # Create history entry
        SalesOrderHistory.objects.create(
            order=order,
            action='REJECTED',
            old_status=old_status,
            new_status=order.status,
            changed_by=request.user if request.user.is_authenticated else None,
            remarks=f'Order rejected: {reason}'
        )
        
        serializer = SalesOrderSerializer(order)
        return Response(serializer.data)


class SalesOrderProcess(APIView):
    """Start processing a sales order"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        qs = SalesOrder.filtered_objects.get_qs(user=request.user)
        qs = apply_company_location_filter(qs, request.user, company_field='company')
        order = get_object_or_404(qs, pk=pk)

        if order.status != 'APPROVED':
            return Response(
                {'error': 'Only APPROVED orders can be processed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = 'PROCESSING'
        order.save()
        
        # Create history entry
        SalesOrderHistory.objects.create(
            order=order,
            action='PROCESSING',
            old_status=old_status,
            new_status=order.status,
            changed_by=request.user if request.user.is_authenticated else None,
            remarks='Processing started'
        )
        
        serializer = SalesOrderSerializer(order)
        return Response(serializer.data)


class SalesOrderInvoice(APIView):
    """Mark sales order as invoiced"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        qs = SalesOrder.filtered_objects.get_qs(user=request.user)
        qs = apply_company_location_filter(qs, request.user, company_field='company')
        order = get_object_or_404(qs, pk=pk)

        if order.status != 'PROCESSING':
            return Response(
                {'error': 'Only PROCESSING orders can be invoiced'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = 'INVOICED'
        order.save()
        
        # Create history entry
        SalesOrderHistory.objects.create(
            order=order,
            action='INVOICED',
            old_status=old_status,
            new_status=order.status,
            changed_by=request.user if request.user.is_authenticated else None,
            remarks='Invoice created'
        )
        
        serializer = SalesOrderSerializer(order)
        return Response(serializer.data)


class SalesOrderDeliver(APIView):
    """Mark sales order as delivered"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, pk):
        qs = SalesOrder.filtered_objects.get_qs(user=request.user)
        qs = apply_company_location_filter(qs, request.user, company_field='company')
        order = get_object_or_404(qs, pk=pk)

        if order.status != 'INVOICED':
            return Response(
                {'error': 'Only INVOICED orders can be marked as delivered'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = 'DELIVERED'
        order.save()
        
        # Create history entry
        SalesOrderHistory.objects.create(
            order=order,
            action='DELIVERED',
            old_status=old_status,
            new_status=order.status,
            changed_by=request.user if request.user.is_authenticated else None,
            remarks='Order delivered'
        )
        
        serializer = SalesOrderSerializer(order)
        return Response(serializer.data)


class GetItemPriceCascade(APIView):
    """Get item price using cascade logic"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        from Masters.models import PriceBook, Retailer, Distributor, Superstockist
        
        item_id = request.query_params.get('item_id')
        customer_id = request.query_params.get('customer_id')
        customer_type = request.query_params.get('customer_type')
        
        if not all([item_id, customer_id, customer_type]):
            return Response(
                {'error': 'item_id, customer_id, and customer_type are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            today = date.today()
            # Base date filter: effective_from <= today AND (effective_to is null OR effective_to >= today)
            date_filter = Q(effective_from__lte=today) & (Q(effective_to__isnull=True) | Q(effective_to__gte=today))
            # Exclude entries whose parent document is CLOSED
            active_doc_filter = ~Q(document__status='CLOSED')

            # Get customer details based on type
            if customer_type == 'RETAILER':
                customer = Retailer.objects.get(pk=customer_id)
                state_id = customer.state_id
                city_id = customer.city_id
                area_id = customer.area_id
            elif customer_type == 'DISTRIBUTOR':
                customer = Distributor.objects.get(pk=customer_id)
                state_id = customer.state_id
                city_id = customer.city_id
                area_id = customer.area_id
            elif customer_type == 'SUPERSTOCKIST':
                customer = Superstockist.objects.get(pk=customer_id)
                state_id = customer.state_id
                city_id = customer.city_id
                area_id = customer.area_id
            else:
                return Response(
                    {'error': 'Invalid customer_type'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Cascade logic: Retailer -> Distributor -> Superstockist -> Area -> City -> State
            price = None
            source = 'NOT_FOUND'
            source_id = None
            
            # Try specific customer
            if customer_type == 'RETAILER':
                price_obj = PriceBook.objects.filter(
                    date_filter, active_doc_filter, item_id=item_id, retailer_id=customer_id, is_active=True, is_deleted=False
                ).first()
                if price_obj:
                    price = price_obj.selling_price
                    source = 'RETAILER'
                    source_id = customer_id
            
            if not price and customer_type in ['RETAILER', 'DISTRIBUTOR']:
                distributor_id = customer.distributor_id if customer_type == 'RETAILER' else customer_id
                if distributor_id:
                    price_obj = PriceBook.objects.filter(
                        date_filter, active_doc_filter, item_id=item_id, distributor_id=distributor_id, is_active=True, is_deleted=False
                    ).first()
                    if price_obj:
                        price = price_obj.selling_price
                        source = 'DISTRIBUTOR'
                        source_id = distributor_id
            
            if not price:
                superstockist_id = None
                if customer_type == 'RETAILER' and hasattr(customer, 'distributor') and customer.distributor:
                    superstockist_id = customer.distributor.superstockist_id
                elif customer_type == 'DISTRIBUTOR':
                    superstockist_id = customer.superstockist_id
                elif customer_type == 'SUPERSTOCKIST':
                    superstockist_id = customer_id
                
                if superstockist_id:
                    price_obj = PriceBook.objects.filter(
                        date_filter, active_doc_filter, item_id=item_id, superstockist_id=superstockist_id, is_active=True, is_deleted=False
                    ).first()
                    if price_obj:
                        price = price_obj.selling_price
                        source = 'SUPERSTOCKIST'
                        source_id = superstockist_id
            
            # Try Area
            if not price and area_id:
                price_obj = PriceBook.objects.filter(
                    date_filter, active_doc_filter, item_id=item_id, area_id=area_id, is_active=True, is_deleted=False
                ).first()
                if price_obj:
                    price = price_obj.selling_price
                    source = 'AREA'
                    source_id = area_id
            
            # Try City
            if not price and city_id:
                price_obj = PriceBook.objects.filter(
                    date_filter, active_doc_filter, item_id=item_id, city_id=city_id, is_active=True, is_deleted=False
                ).first()
                if price_obj:
                    price = price_obj.selling_price
                    source = 'CITY'
                    source_id = city_id
            
            # Try State
            if not price and state_id:
                price_obj = PriceBook.objects.filter(
                    date_filter, active_doc_filter, item_id=item_id, state_id=state_id, is_active=True, is_deleted=False
                ).first()
                if price_obj:
                    price = price_obj.selling_price
                    source = 'STATE'
                    source_id = state_id

            # Try BASE price (final fallback)
            if not price:
                price_obj = PriceBook.objects.filter(
                    date_filter, active_doc_filter, item_id=item_id, price_type='BASE', is_active=True, is_deleted=False
                ).first()
                if price_obj:
                    price = price_obj.selling_price
                    source = 'BASE'
                    source_id = None
            
            return Response({
                'pb_rate': float(price) if price else None,
                'source': source,
                'source_id': str(source_id) if source_id else None,
                'found': price is not None
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class GenerateSalesOrderNumber(APIView):
    """
    Generate next sales order document number
    
    GET /api/sales/orders/generate-document-number/
    Returns:
    {
        "order_number": "SO-25-26-1",
        "financial_year": "25-26"
    }
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from django.utils import timezone
        import re
        
        today = timezone.now().date()
        
        # Determine financial year (April 1 to March 31)
        if today.month >= 4:  # April onwards
            fy_start = today.year
            fy_end = today.year + 1
        else:  # January to March
            fy_start = today.year - 1
            fy_end = today.year
        
        # Format: SO-25-26-1 (where 25-26 is FY 2025-26)
        fy_suffix = f"{str(fy_start)[-2:]}-{str(fy_end)[-2:]}"
        prefix = f"SO-{fy_suffix}"
        
        # Get all matching order numbers and compute the max numeric suffix
        order_numbers = SalesOrder.objects.filter(
            order_number__startswith=prefix
        ).values_list('order_number', flat=True)

        max_suffix = 0
        for on in order_numbers:
            match = re.search(r'-(\d+)$', on or '')
            if match:
                try:
                    max_suffix = max(max_suffix, int(match.group(1)))
                except ValueError:
                    continue

        order_number = f"{prefix}-{max_suffix + 1}"
        
        return Response({
            'order_number': order_number,
            'financial_year': fy_suffix
        }, status=status.HTTP_200_OK)


class SalesOrderHistoryList(generics.ListAPIView):
    """List sales order history"""
    permission_classes = [permissions.AllowAny]
    serializer_class = SalesOrderHistorySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    ordering = ['-changed_at']
    
    def get_queryset(self):
        order_id = self.kwargs.get('order_id')
        return SalesOrderHistory.objects.filter(order_id=order_id).select_related(
            'order'
        )


class GetCustomerPendingSalesOrdersView(APIView):
    """Get pending sales orders for a customer"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        """
        Fetch pending sales orders for a customer
        Query params: customer_type, customer_id
        Returns orders that are not yet fully invoiced
        """
        from django.utils import timezone
        from django.db.models import Sum, Q
        from decimal import Decimal
        
        customer_type = request.query_params.get('customer_type')
        customer_id = request.query_params.get('customer_id')
        
        if not customer_type or not customer_id:
            return Response(
                {'error': 'Both customer_type and customer_id parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build filter based on customer type
        filter_kwargs = {
            'status__in': ['DRAFT', 'CONFIRMED', 'PARTIALLY_DISPATCHED', 'DISPATCHED'],
            'authorized_status__in': [1, 2],  # PENDING or APPROVED, exclude REJECTED (3)
        }
        
        # Map customer type to filter field
        if customer_type == 'RETAILER':
            filter_kwargs['retailer_id'] = customer_id
        elif customer_type == 'DISTRIBUTOR':
            filter_kwargs['distributor_id'] = customer_id
        elif customer_type == 'SUPERSTOCKIST':
            filter_kwargs['superstockist_id'] = customer_id
        else:
            return Response(
                {'error': f'Invalid customer_type: {customer_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Fetch pending orders filtered by user's company access
            orders = SalesOrder.objects.filter(**filter_kwargs).select_related(
                'company'
            ).order_by('-order_date')
            orders = apply_company_location_filter(orders, request.user, company_field='company')
            
            # Serialize data with calculated days and uninvoiced amount
            data = []
            today = timezone.now().date()
            total_order_value = Decimal('0')
            
            for order in orders:
                # Calculate invoiced amount for this order
                from Invoice.models import Invoice
                invoiced_amount = Invoice.objects.filter(
                    sales_order=order,
                    status__in=['CONFIRMED', 'PAID', 'PARTIALLY_PAID']
                ).aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
                
                # Calculate uninvoiced amount
                uninvoiced_amount = order.grand_total - invoiced_amount
                
                # Only include if there's uninvoiced amount
                if uninvoiced_amount > 0:
                    days_diff = (today - order.order_date).days
                    data.append({
                        'id': str(order.id),
                        'order_number': order.order_number,
                        'order_date': order.order_date.strftime('%d-%m-%Y'),
                        'grand_total': float(order.grand_total),
                        'invoiced_amount': float(invoiced_amount),
                        'uninvoiced_amount': float(uninvoiced_amount),
                        'status': order.status,
                        'days': days_diff
                    })
                    total_order_value += uninvoiced_amount
            
            # Get customer credit limit and pending invoice balance
            credit_limit = Decimal('0')
            if customer_type == 'RETAILER':
                from Masters.models import Retailer
                customer = Retailer.objects.filter(id=customer_id).first()
                if customer:
                    credit_limit = customer.credit_limit
            elif customer_type == 'DISTRIBUTOR':
                from Masters.models import Distributor
                customer = Distributor.objects.filter(id=customer_id).first()
                if customer:
                    credit_limit = customer.credit_limit
            elif customer_type == 'SUPERSTOCKIST':
                from Masters.models import Superstockist
                customer = Superstockist.objects.filter(id=customer_id).first()
                if customer:
                    credit_limit = customer.credit_limit
            
            # Get pending invoice balance
            from Invoice.models import Invoice
            pending_invoice_balance = Invoice.objects.filter(
                **{f'sales_order__{customer_type.lower()}_id': customer_id},
                status__in=['CONFIRMED', 'PAID', 'PARTIALLY_PAID'],
                balance_amount__gt=0
            ).aggregate(total=Sum('balance_amount'))['total'] or Decimal('0')
            
            # Calculate available credit
            available_credit = credit_limit - total_order_value - pending_invoice_balance
            
            return Response({
                'count': len(data),
                'results': data,
                'credit_summary': {
                    'credit_limit': float(credit_limit),
                    'existing_order_value': float(total_order_value),
                    'pending_invoice_balance': float(pending_invoice_balance),
                    'available_credit': float(available_credit)
                }
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching pending sales orders: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class GetCustomerFrequentItemsView(APIView):
    """
    Get customer's most frequently ordered items.

    Query params:
    - customer_type: RETAILER|DISTRIBUTOR|SUPERSTOCKIST
    - customer_id: UUID
    - months: optional (default 12)
    - limit: optional (default 10)
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        from django.db.models import Sum, Count, Max
        from django.utils import timezone

        customer_type = request.query_params.get('customer_type')
        customer_id = request.query_params.get('customer_id')
        company_id = request.query_params.get('company_id')
        months = int(request.query_params.get('months', 12) or 12)
        limit = int(request.query_params.get('limit', 10) or 10)
        limit = max(1, min(limit, 100))

        if not customer_type or not customer_id:
            return Response(
                {'error': 'Both customer_type and customer_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        customer_type = customer_type.upper()
        if customer_type not in ['RETAILER', 'DISTRIBUTOR', 'SUPERSTOCKIST']:
            return Response(
                {'error': f'Invalid customer_type: {customer_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from_date = timezone.now().date() - timedelta(days=max(months, 1) * 30)
        order_filter = {
            'order_date__gte': from_date,
            # Include finalized pipeline statuses; exclude draft/cancelled/rejected-like
            'status__in': [
                'CONFIRMED',
                'PARTIALLY_DISPATCHED',
                'DISPATCHED',
                'PARTIALLY_INVOICED',
                'INVOICED',
                'DELIVERED',
            ],
        }

        if customer_type == 'RETAILER':
            order_filter['retailer_id'] = customer_id
        elif customer_type == 'DISTRIBUTOR':
            order_filter['distributor_id'] = customer_id
        else:
            order_filter['superstockist_id'] = customer_id

        orders_qs = SalesOrder.objects.filter(**order_filter)
        orders_qs = apply_company_location_filter(orders_qs, request.user, company_field='company')

        if GeneralSettings.is_company_scoped_item_enforcement_enabled():
            if company_id:
                orders_qs = orders_qs.filter(company_id=company_id)
            else:
                orders_qs = orders_qs.none()

        line_items_qs = SalesOrderItem.objects.filter(order__in=orders_qs, is_scheme_item=False)

        # When company-scoped item enforcement is enabled, restrict frequent items
        # by the selected header company at item master level as well.
        if GeneralSettings.is_company_scoped_item_enforcement_enabled():
            line_items_qs = line_items_qs.filter(item__company_id=company_id)

        frequent_items = (
            line_items_qs
            .values(
                'item_id',
                'item__code',
                'item__name',
                'item__hsn_code',
                'item__base_uom__name',
                'category_id',
                'category__name',
                'item__company_id',
                'item__company__name',
            )
            .annotate(
                total_ordered_qty=Sum('quantity'),
                order_count=Count('order_id', distinct=True),
                last_order_date=Max('order__order_date'),
            )
            .order_by('-total_ordered_qty', '-last_order_date')[:limit]
        )

        results = [
            {
                'item': row['item_id'],
                'item_code': row['item__code'],
                'item_name': row['item__name'],
                'hsn_code': row['item__hsn_code'],
                'uom_name': row['item__base_uom__name'],
                'category': row['category_id'],
                'category_name': row['category__name'],
                'company': row['item__company_id'],
                'company_name': row['item__company__name'],
                'total_ordered_qty': float(row['total_ordered_qty'] or 0),
                'order_count': int(row['order_count'] or 0),
                'last_order_date': row['last_order_date'],
            }
            for row in frequent_items
        ]

        return Response(
            {
                'count': len(results),
                'months': months,
                'results': results,
            },
            status=status.HTTP_200_OK
        )

class GetApplicableSchemesView(APIView):
    """
    View to get applicable schemes for a given order.
    
    Expected POST data:
    {
        "customer_type": "superstockist|distributor|retailer",
        "customer_id": "uuid",
        "company_id": "uuid",
        "location_id": "uuid (optional)",
        "items": [
            {
                "item_id": "uuid",
                "quantity": 10,
                "unit": "BOX",
                "item_amount": 5000
            }
        ],
        "order_date": "2024-01-15"
    }
    
    Returns:
    {
        "applicable_schemes": [
            {
                "id": "uuid",
                "code": "SCHEME001",
                "name": "10% Discount on Orders",
                "scheme_type": "PERCENTAGE",
                "status": "ACTIVE",
                ...
            }
        ]
    }
    """
    permission_classes = [permissions.AllowAny]
    serializer_class = SchemeMiniSerializer
    
    def post(self, request, *args, **kwargs):
        """Get applicable schemes for order"""
        try:
            customer_type = (request.data.get('customer_type') or '').upper()
            customer_id = request.data.get('customer_id')
            # company_id = request.data.get('company_id')
            location_id = request.data.get('location_id')
            items = request.data.get('items', [])
            order_date = request.data.get('order_date', str(date.today()))
            
            # Parse order date
            try:
                order_date = datetime.strptime(order_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                order_date = date.today()
            
            # if not company_id:
            #     return Response(
            #         {'error': 'Company is required to fetch schemes.'},
            #         status=status.HTTP_400_BAD_REQUEST
            #     )

            # Get all active schemes for the company
            schemes = Scheme.objects.filter(
                # company_id=company_id,
                status='ACTIVE',
                is_deleted=False,
                authorized_status=2,
                effective_from__lte=order_date
            ).filter(
                models.Q(effective_to__isnull=True) | models.Q(effective_to__gte=order_date)
            ).prefetch_related(
                'applicability',
                'items',
                'conditions',
                'benefits'
            ).order_by('priority', 'code')

            # Resolve location_id to state/city/area if it's a Location UUID
            location_state_id = None
            location_city_id = None
            location_area_id = None
            if location_id:
                try:
                    from Masters.models import Location
                    location_obj = Location.objects.filter(id=location_id).first()
                    if location_obj:
                        location_state_id = getattr(location_obj, 'state_id', None)
                        location_city_id = getattr(location_obj, 'city_id', None)
                        location_area_id = getattr(location_obj, 'area_id', None)
                except Exception:
                    # If Location lookup fails, fallback to raw id matching
                    location_state_id = location_id
                    location_city_id = location_id
                    location_area_id = location_id

            # Build a lightweight order-like object for preview + validation
            class _TempItem:
                def __init__(self, item_id, quantity, line_total, category_id=None, item_name=None):
                    self.item_id = item_id
                    self.quantity = Decimal(str(quantity or 0))
                    self.line_total = Decimal(str(line_total or 0))
                    self.category_id = category_id
                    self.item_name = item_name
                    self.item = None

            class _TempOrder:
                def __init__(self):
                    # self.company = company_id
                    self.billing_state_id = None
                    self.billing_city_id = None
                    self.billing_area_id = None
                    self.customer_type = customer_type
                    self.retailer_id = None
                    self.distributor_id = None
                    self.superstockist_id = None

            temp_order = _TempOrder()

            # Allow location_id to match any of state/city/area applicability
            if location_state_id or location_city_id or location_area_id:
                temp_order.billing_state_id = location_state_id
                temp_order.billing_city_id = location_city_id
                temp_order.billing_area_id = location_area_id

            if customer_type == 'RETAILER':
                temp_order.retailer_id = customer_id
            elif customer_type == 'DISTRIBUTOR':
                temp_order.distributor_id = customer_id
            elif customer_type == 'SUPERSTOCKIST':
                temp_order.superstockist_id = customer_id

            # Pre-fetch item names for temp items
            from Masters.models import Item as MasterItem
            item_ids = [item.get('item_id') for item in items if item.get('item_id')]
            item_name_map = {}
            if item_ids:
                item_name_map = {
                    str(k): v for k, v in
                    MasterItem.objects.filter(id__in=item_ids).values_list('id', 'name')
                }

            temp_items = [
                _TempItem(
                    item.get('item_id'),
                    item.get('quantity') or 0,
                    item.get('item_amount') or 0,
                    item.get('category_id'),
                    item.get('item_name') or (item_name_map.get(str(item.get('item_id')), '') if item.get('item_id') else '')
                )
                for item in items
            ]

            from Masters.scheme_engine import SchemeEngine
            engine = SchemeEngine()
            preview_results = []
            debug_mode = str(request.query_params.get('debug', '')).lower() in ('1', 'true', 'yes')
            debug_reasons = [] if debug_mode else None

            def matches_applicability(scheme_obj):
                # Geographic applicability
                applicabilities = scheme_obj.applicability.all()
                if applicabilities.exists():
                    geo_match = False
                    for applicability in applicabilities:
                        if not applicability.state_id and not applicability.city_id and not applicability.area_id:
                            geo_match = True
                            break
                        if location_id and (
                            applicability.state_id in (location_state_id, location_id)
                            or applicability.city_id in (location_city_id, location_id)
                            or applicability.area_id in (location_area_id, location_id)
                        ):
                            geo_match = True
                            break
                    if not geo_match:
                        return False

                # Channel applicability
                if applicabilities.exists():
                    channel_match = False
                    for applicability in applicabilities:
                        if applicability.customer_type == 'ALL':
                            channel_match = True
                            break
                        if applicability.customer_type == customer_type:
                            if customer_type == 'RETAILER':
                                channel_match = applicability.retailer_id in (None, customer_id)
                            elif customer_type == 'DISTRIBUTOR':
                                channel_match = applicability.distributor_id in (None, customer_id)
                            elif customer_type == 'SUPERSTOCKIST':
                                channel_match = applicability.superstockist_id in (None, customer_id)
                            else:
                                channel_match = True
                            if channel_match:
                                break
                    if not channel_match:
                        return False

                # Item applicability
                scheme_items = scheme_obj.items.all()
                if scheme_items.exists():
                    if scheme_items.filter(include_all_items=True).exists():
                        return True
                    order_item_ids = {item.item_id for item in temp_items if item.item_id}
                    scheme_item_ids = set(scheme_items.filter(item_id__isnull=False).values_list('item_id', flat=True))
                    if order_item_ids & scheme_item_ids:
                        return True
                    scheme_category_ids = set(scheme_items.filter(category_id__isnull=False).values_list('category_id', flat=True))
                    order_category_ids = {item.category_id for item in temp_items if item.category_id}
                    if order_category_ids & scheme_category_ids:
                        return True
                    return False

                return True

            for scheme in schemes:
                if not matches_applicability(scheme):
                    if debug_mode:
                        debug_reasons.append({
                            'scheme_id': str(scheme.id),
                            'code': scheme.code,
                            'name': scheme.name,
                            'reason': 'applicability_mismatch'
                        })
                    continue

                is_valid, matching_conditions, failed_conditions = engine.validate_conditions(scheme, temp_order, temp_items)
                if not is_valid:
                    if debug_mode:
                        debug_reasons.append({
                            'scheme_id': str(scheme.id),
                            'code': scheme.code,
                            'name': scheme.name,
                            'reason': 'condition_failed',
                            'failed_conditions': [c.condition_type for c in failed_conditions],
                            'matched_conditions': [c.condition_type for c in matching_conditions],
                        })
                    continue

                benefits = engine.calculate_benefits(scheme, temp_order, temp_items)
                cart_total = sum(item.line_total for item in temp_items) or Decimal('0')
                discount_amount = Decimal(str(benefits.get('discount_amount', 0) or 0))
                if cart_total > 0 and discount_amount > cart_total:
                    if debug_mode:
                        debug_reasons.append({
                            'scheme_id': str(scheme.id),
                            'code': scheme.code,
                            'name': scheme.name,
                            'reason': 'discount_exceeds_cart',
                            'discount_amount': str(discount_amount),
                            'cart_total': str(cart_total),
                        })
                    continue

                exceeds_benefit_base = False
                for detail in benefits.get('benefit_details', []):
                    if isinstance(detail, dict) and detail.get('type') in ('DISCOUNT_AMOUNT', 'DISCOUNT_PERCENTAGE'):
                        base_amount = Decimal(str(detail.get('base_amount', 0) or 0))
                        amount = Decimal(str(detail.get('amount', 0) or 0))
                        if base_amount > 0 and amount > base_amount:
                            exceeds_benefit_base = True
                            break
                if exceeds_benefit_base:
                    if debug_mode:
                        debug_reasons.append({
                            'scheme_id': str(scheme.id),
                            'code': scheme.code,
                            'name': scheme.name,
                            'reason': 'discount_exceeds_base',
                        })
                    continue

                preview_results.append({
                    'id': str(scheme.id),
                    'code': scheme.code,
                    'name': scheme.name,
                    'priority': scheme.priority,
                    'scheme_type': scheme.scheme_type,
                    'scheme_type_display': scheme.get_scheme_type_display(),
                    'preview_discount_amount': float(benefits.get('discount_amount', 0) or 0),
                    'preview_total_discount': float(benefits.get('total_discount', 0) or 0),
                    'preview_free_items': benefits.get('free_items', []),
                    'preview_benefit_details': SalesOrderAvailableSchemesView._serialize_benefit_details(
                        benefits.get('benefit_details', [])
                    ),
                })
            
            response_payload = {
                'applicable_schemes': preview_results,
                'count': len(preview_results)
            }
            if debug_mode:
                response_payload['debug'] = debug_reasons
                response_payload['debug_summary'] = {
                    'total_active_schemes': schemes.count(),
                    'matched': len(preview_results),
                    'filtered': len(debug_reasons),
                }
            return Response(response_payload, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error getting applicable schemes: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

class SalesOrderAvailableSchemesView(APIView):
    """
    Get applicable schemes for a saved sales order with preview benefits.
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request, pk, *args, **kwargs):
        try:
            order = SalesOrder.objects.prefetch_related('items').get(id=pk)

            from Masters.scheme_engine import SchemeEngine
            engine = SchemeEngine()
            order_items = list(order.items.all())
            applicable = engine.get_applicable_schemes(order, order_date=order.order_date)

            results = []
            for scheme in applicable:
                is_valid, _, _ = engine.validate_conditions(scheme, order, order_items)
                if not is_valid:
                    continue

                benefits = engine.calculate_benefits(scheme, order, order_items)
                cart_total = sum(item.line_total for item in order_items) or Decimal('0')
                discount_amount = Decimal(str(benefits.get('discount_amount', 0) or 0))
                if cart_total > 0 and discount_amount > cart_total:
                    continue

                exceeds_benefit_base = False
                for detail in benefits.get('benefit_details', []):
                    if isinstance(detail, dict) and detail.get('type') in ('DISCOUNT_AMOUNT', 'DISCOUNT_PERCENTAGE'):
                        base_amount = Decimal(str(detail.get('base_amount', 0) or 0))
                        amount = Decimal(str(detail.get('amount', 0) or 0))
                        if base_amount > 0 and amount > base_amount:
                            exceeds_benefit_base = True
                            break
                if exceeds_benefit_base:
                    continue

                results.append({
                    'id': str(scheme.id),
                    'code': scheme.code,
                    'name': scheme.name,
                    'priority': scheme.priority,
                    'scheme_type': scheme.scheme_type,
                    'scheme_type_display': scheme.get_scheme_type_display(),
                    'preview_discount_amount': float(benefits.get('discount_amount', 0) or 0),
                    'preview_total_discount': float(benefits.get('total_discount', 0) or 0),
                    'preview_free_items': benefits.get('free_items', []),
                    'preview_benefit_details': self._serialize_benefit_details(benefits.get('benefit_details', [])),
                })

            results.sort(key=lambda x: (x.get('priority', 0), x.get('code', '')))

            return Response({
                'applicable_schemes': results,
                'count': len(results)
            }, status=status.HTTP_200_OK)

        except SalesOrder.DoesNotExist:
            return Response(
                {'error': 'Sales order not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error getting applicable schemes: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @staticmethod
    def _serialize_benefit_details(details):
        serialized = []
        for detail in details:
            if isinstance(detail, dict):
                cleaned = {}
                for key, value in detail.items():
                    if isinstance(value, (int, float, str, bool)) or value is None:
                        cleaned[key] = value
                    elif hasattr(value, 'quantize'):
                        cleaned[key] = float(value)
                    elif isinstance(value, list):
                        cleaned[key] = value
                    else:
                        cleaned[key] = str(value)
                serialized.append(cleaned)
        return serialized


class ApplySchemesToOrderView(APIView):
    """
    Apply selected schemes to an existing sales order.
    
    Expected POST data:
    {
        "scheme_ids": ["uuid1", "uuid2"],
        "auto_recalculate": true
    }
    
    Updates the order with:
    - Applied schemes in SalesOrderScheme
    - Item-level discounts in SalesOrderItemScheme
    - Updated order totals
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, pk, *args, **kwargs):
        """Apply schemes to existing order"""
        try:
            order = SalesOrder.objects.get(id=pk)
            scheme_ids = request.data.get('scheme_ids', [])
            auto_recalculate = request.data.get('auto_recalculate', True)
            
            if not scheme_ids:
                return Response(
                    {'error': 'No schemes provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if (not GeneralSettings.is_allow_multiple_schemes_enabled()) and len(scheme_ids) > 1:
                return Response(
                    {'error': 'Only one scheme can be selected per order.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get selected schemes
            schemes = Scheme.objects.filter(
                id__in=scheme_ids,
                is_deleted=False,
                status='ACTIVE',
                authorized_status=2
            ).order_by('priority', 'code')
            
            if not schemes.exists():
                return Response(
                    {'error': 'No valid schemes found'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Apply schemes
            from Masters.scheme_engine import SchemeEngine
            engine = SchemeEngine()
            result = engine.apply_schemes(order, list(schemes), request.user)
            
            if auto_recalculate and result['success']:
                order.calculate_totals()
                order.save()
            
            return Response(result, status=status.HTTP_200_OK)
            
        except SalesOrder.DoesNotExist:
            return Response(
                {'error': 'Sales order not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error applying schemes: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class SalesOrderStatusCountView(generics.ListAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = SalesOrderStatusCountSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = SalesOrderFilter

    def get_queryset(self):
        queryset = SalesOrder.filtered_objects.get_qs(
            user=self.request.user
        )

        queryset = apply_company_location_filter(
            queryset,
            self.request.user,
            company_field='company'
        )

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        status_counts = queryset.values('status').annotate(
            count=Count('id')
        )

        data = {
            'DRAFT': 0,
            'PENDING': 0,
            'CONFIRMED': 0,
            'PARTIALLY_DISPATCHED': 0,
            'DISPATCHED': 0,
            'PARTIALLY_INVOICED': 0,
            'INVOICED': 0,
            'DELIVERED': 0,
            'CANCELLED': 0,
        }

        for item in status_counts:
            if item['status'] in data:
                data[item['status']] = item['count']

        data['total'] = sum(data.values())

        serializer = self.get_serializer(data)
        return Response(serializer.data)


class SalesOrderWeeklyCustomerTypeCountView(APIView):
    """
    Get sales order counts for last 4 weeks.
    
    Query Parameters:
    - customer_type: Filter by customer type (RETAILER, DISTRIBUTOR, SUPERSTOCKIST)
    - company_id: Filter by company (optional)
    
    Returns counts for each of the last 4 weeks:
    - Week 4: Last 7 days (today to today - 6 days)
    - Week 3: Previous 7 days (today - 7 to today - 13 days)
    - Week 2: Previous 7 days (today - 14 to today - 20 days)
    - Week 1: Previous 7 days (today - 21 to today - 27 days)
    
    Sample response:
    {
        "date_range": "Feb 13 to Mar 12, 2026",
        "customer_type": "RETAILER",
        "weeks": {
            "Week 4": {
                "start_date": "2026-03-06",
                "end_date": "2026-03-12",
                "count": 10
            },
            "Week 3": {...},
            "Week 2": {...},
            "Week 1": {...}
        },
        "total": 40
    }
    """
    permission_classes = [permissions.AllowAny]
    
    def get(self, request, *args, **kwargs):
        try:
            from django.utils import timezone
            
            today = timezone.now().date()
            
            # Calculate week ranges backward from today
            # Week 4: today to today - 6 days
            # Week 3: today - 7 to today - 13 days
            # Week 2: today - 14 to today - 20 days
            # Week 1: today - 21 to today - 27 days
            week_ranges = {
                'Week 4': (today - timedelta(days=6), today),
                'Week 3': (today - timedelta(days=13), today - timedelta(days=7)),
                'Week 2': (today - timedelta(days=20), today - timedelta(days=14)),
                'Week 1': (today - timedelta(days=27), today - timedelta(days=21)),
            }
            
            # Get base queryset
            queryset = SalesOrder.filtered_objects.get_qs(
                user=request.user
            )
            queryset = apply_company_location_filter(
                queryset, 
                request.user, 
                company_field='company'
            )
            
            # Apply company filter if provided
            company_id = request.query_params.get('company_id')
            if company_id:
                queryset = queryset.filter(company_id=company_id)
            
            # Apply customer_type filter
            customer_type = request.query_params.get('customer_type')
            if customer_type and customer_type.upper() in ['RETAILER', 'DISTRIBUTOR', 'SUPERSTOCKIST']:
                queryset = queryset.filter(customer_type=customer_type.upper())
            elif customer_type:
                return Response(
                    {'error': f'Invalid customer_type: {customer_type}. Must be RETAILER, DISTRIBUTOR, or SUPERSTOCKIST'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Build response
            range_start = today - timedelta(days=27)
            range_end = today
            response_data = {
                'date_range': f"{range_start.strftime('%b %d')} to {range_end.strftime('%b %d, %Y')}",
                'customer_type': customer_type.upper() if customer_type else 'ALL',
                'weeks': {},
                'total': 0
            }
            
            # Process each week
            for week_name, (week_start, week_end) in week_ranges.items():
                week_count = queryset.filter(
                    order_date__gte=week_start,
                    order_date__lte=week_end
                ).count()
                
                response_data['weeks'][week_name] = {
                    'start_date': week_start.isoformat(),
                    'end_date': week_end.isoformat(),
                    'count': week_count
                }
                response_data['total'] += week_count
            
            return Response(response_data, status=status.HTTP_200_OK)
            
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error fetching weekly customer type counts: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OrderFulfilmentPercentageView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        queryset = SalesOrder.filtered_objects.get_qs(user=request.user).exclude(status__in=['DRAFT', 'CANCELLED'])
        queryset = apply_company_location_filter(queryset, request.user, company_field='company')
        total = queryset.count()
        if total == 0:
            return Response({'total_orders': 0, 'dispatched_orders': 0, 'fulfilment_percentage': 0})
        dispatched = queryset.filter(status='DELIVERED').count()
        percentage = round((dispatched / total) * 100, 2)
        return Response({'total_orders': total, 'dispatched_orders': dispatched, 'fulfilment_percentage': percentage})


class SalesOrderCustomerDropdown(APIView):
    """
    Returns channel partner list for sales order filtering based on logged-in user.

    GET /api/sales/orders/customer-dropdown/?customer_type=RETAILER
    GET /api/sales/orders/customer-dropdown/?customer_type=DISTRIBUTOR
    GET /api/sales/orders/customer-dropdown/?customer_type=SUPERSTOCKIST

    - Staff user: sees all active channel partners of that type
    - Distributor user: customer_type=RETAILER returns retailers under their distributor
    - Superstockist user: customer_type=DISTRIBUTOR returns distributors under them,
                          customer_type=RETAILER returns retailers under those distributors
    - Retailer user: only sees themselves
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from Masters.models import Retailer, Distributor, Superstockist
        from utils import apply_channel_partner_company_location_filter

        customer_type = request.query_params.get('customer_type', '').upper()
        search = request.query_params.get('search', '')
        user = request.user

        if customer_type == 'RETAILER':
            qs = Retailer.filtered_objects.get_qs(user=user, is_active=True)
            qs = apply_channel_partner_company_location_filter(
                qs, user, company_field='company', state_field='state',
                city_field='city', coverage_relation='locations'
            )
        elif customer_type == 'DISTRIBUTOR':
            qs = Distributor.filtered_objects.get_qs(user=user, is_active=True)
            qs = apply_channel_partner_company_location_filter(
                qs, user, company_field='company', state_field='state',
                city_field='city', coverage_relation='locations'
            )
        elif customer_type == 'SUPERSTOCKIST':
            qs = Superstockist.filtered_objects.get_qs(user=user, is_active=True)
            qs = apply_channel_partner_company_location_filter(
                qs, user, company_field='company', state_field='state',
                city_field='city', coverage_relation='locations'
            )
        else:
            return Response(
                {'error': 'customer_type is required (RETAILER, DISTRIBUTOR, SUPERSTOCKIST)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(code__icontains=search))

        results = list(qs.distinct().values('id', 'code', 'name').order_by('name'))
        return Response({'count': len(results), 'results': results}, status=status.HTTP_200_OK)


class SalesOrderReportFilter(FilterSet):
    """
    Advanced filter for Sales Order Report with comprehensive filtering options
    """
    # Date Filters
    from_date = DateFilter(field_name='order_date', lookup_expr='gte', label='From Date')
    to_date = DateFilter(field_name='order_date', lookup_expr='lte', label='To Date')
    
    # Quick Date Filters (handled in view)
    date_preset = CharFilter(method='filter_date_preset', label='Date Preset (today/this_week/this_month/this_year)')
    
    # Customer Filters
    customer_type = ChoiceFilter(
        field_name='customer_type',
        choices=[
            ('RETAILER', 'Retailer'),
            ('DISTRIBUTOR', 'Distributor'),
            ('SUPERSTOCKIST', 'Superstockist'),
        ],
        label='Customer Type'
    )
    customer_id = CharFilter(method='filter_customer', label='Customer ID')
    
    # Authorization Filters
    authorization_status = ChoiceFilter(
        field_name='authorized_status',
        choices=[
            ('1', 'Pending'),
            ('2', 'Approved'),
            ('3', 'Rejected'),
        ],
        label='Authorization Status'
    )
    
    # Location Filters (Hierarchical)
    country = UUIDFilter(method='filter_country', label='Country')
    state = UUIDFilter(method='filter_state', label='State')
    district = UUIDFilter(method='filter_district', label='District')
    mandal = UUIDFilter(method='filter_mandal', label='Mandal')
    city = UUIDFilter(method='filter_city', label='City')
    area = UUIDFilter(method='filter_area', label='Area')
    
    # Agent Filter
    agent = UUIDFilter(field_name='distributor__agent_id', label='Agent')

    # Order & Product Filters
    order_status = ChoiceFilter(
        field_name='status',
        choices=[
            ('DRAFT', 'Draft'),
            ('PENDING', 'Pending'),
            ('CONFIRMED', 'Confirmed'),
            ('PARTIALLY_DISPATCHED', 'Partially Dispatched'),
            ('DISPATCHED', 'Dispatched'),
            ('PARTIALLY_INVOICED', 'Partially Invoiced'),
            ('INVOICED', 'Invoiced'),
            ('DELIVERED', 'Delivered'),
            ('CANCELLED', 'Cancelled'),
        ],
        label='Order Status'
    )
    product_id = UUIDFilter(method='filter_product', label='Product ID')
    
    class Meta:
        model = SalesOrder
        fields = [
            'from_date', 'to_date', 'date_preset',
            'customer_type', 'customer_id',
            'authorization_status',
            'agent',
            'country', 'state', 'district', 'mandal', 'city', 'area',
            'order_status', 'product_id'
        ]
    
    def filter_date_preset(self, queryset, name, value):
        """Handle quick date filter presets"""
        today = date.today()
        
        if value == 'today':
            return queryset.filter(order_date=today)
        elif value == 'this_week':
            start_of_week = today - timedelta(days=today.weekday())
            return queryset.filter(order_date__gte=start_of_week, order_date__lte=today)
        elif value == 'this_month':
            start_of_month = today.replace(day=1)
            return queryset.filter(order_date__gte=start_of_month, order_date__lte=today)
        elif value == 'this_year':
            start_of_year = today.replace(month=1, day=1)
            return queryset.filter(order_date__gte=start_of_year, order_date__lte=today)
        
        return queryset
    
    def filter_customer(self, queryset, name, value):
        """Filter by specific customer ID across all customer types"""
        return queryset.filter(
            Q(retailer_id=value) |
            Q(distributor_id=value) |
            Q(superstockist_id=value)
        )
    
    def filter_country(self, queryset, name, value):
        """Filter by country (through state relationship)"""
        return queryset.filter(
            Q(billing_state__country_id=value) |
            Q(shipping_state__country_id=value)
        )
    
    def filter_state(self, queryset, name, value):
        """Filter by state"""
        return queryset.filter(
            Q(billing_state_id=value) |
            Q(shipping_state_id=value)
        )
    
    def filter_district(self, queryset, name, value):
        """Filter by district (through city relationship)"""
        return queryset.filter(
            Q(billing_city__district_id=value) |
            Q(shipping_city__district_id=value)
        )
    
    def filter_mandal(self, queryset, name, value):
        """Filter by mandal (through city relationship)"""
        return queryset.filter(
            Q(billing_city__mandal_id=value) |
            Q(shipping_city__mandal_id=value)
        )
    
    def filter_city(self, queryset, name, value):
        """Filter by city"""
        return queryset.filter(
            Q(billing_city_id=value) |
            Q(shipping_city_id=value)
        )
    
    def filter_area(self, queryset, name, value):
        """Filter by area"""
        return queryset.filter(
            Q(billing_area_id=value) |
            Q(shipping_area_id=value)
        )
    
    def filter_product(self, queryset, name, value):
        """Filter orders that contain a specific product"""
        return queryset.filter(items__item_id=value).distinct()


class SalesOrderReportView(generics.ListAPIView):
    """
    Product-wise Sales Order Report API
    Each row represents ONE product from an order
    """
    
    permission_classes = [GetPermission('Sales.view_salesorder')]
    serializer_class = SalesOrderReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = SalesOrderReportFilter
    
    search_fields = [
        'order_number',
        'code',
        'retailer__name',
        'retailer__code',
        'distributor__name',
        'distributor__code',
        'superstockist__name',
        'superstockist__code',
        'items__item__name',
        'items__item__code',
        'billing_city__name',
        'billing_state__name',
        'billing_area__name',
        'shipping_city__name',
        'shipping_state__name',
        'shipping_area__name',
    ]
    
    ordering_fields = [
        'order_date',
        'order_number',
        'grand_total',
        'status',
        'authorized_status',
    ]
    
    ordering = ['-order_date', 'order_number']
    
    def get_queryset(self):
        queryset = SalesOrder.filtered_objects.get_qs(
            user=self.request.user
        ).select_related(
            'billing_state',
            'billing_state__country',
            'billing_city',
            'billing_city__district',
            'billing_city__mandal',
            'billing_area',
            'shipping_state',
            'shipping_city',
            'shipping_area',
            'retailer',
            'distributor',
            'distributor__agent',
            'superstockist',
            'company'
        ).prefetch_related(
            'items',
            'items__item'
        )
        
        queryset = apply_company_location_filter(queryset, self.request.user, company_field='company')
        return queryset
    
    def _convert_to_product_rows(self, orders):
        """
        Convert orders to product-wise rows
        Each product in an order becomes a separate row
        """
        product_rows = []
        
        for order in orders:
            auth_status_map = {1: 'Pending', 2: 'Approved', 3: 'Rejected', None: 'Not Submitted'}
            auth_status = auth_status_map.get(order.authorized_status, 'Unknown')
            
            country = order.billing_state.country.name if order.billing_state and order.billing_state.country else None
            state = order.billing_state.name if order.billing_state else ''
            state_id = str(order.billing_state.id) if order.billing_state else None
            district = order.billing_city.district.name if order.billing_city and order.billing_city.district else None
            mandal = order.billing_city.mandal.name if order.billing_city and order.billing_city.mandal else None
            city = order.billing_city.name if order.billing_city else ''
            city_id = str(order.billing_city.id) if order.billing_city else None
            area = order.billing_area.name if order.billing_area else None
            area_id = str(order.billing_area.id) if order.billing_area else None
            
            for item in order.items.all():
                product_rows.append({
                    'order_id': str(order.id),
                    'order_number': order.order_number,
                    'order_date': order.order_date,
                    'customer_name': order.get_customer_name(),
                    'customer_type': order.get_customer_type_display(),
                    'customer_type_code': order.customer_type,
                    'product_id': str(item.item.id),
                    'product_code': item.item.code,
                    'product_name': item.item.name,
                    'quantity': item.quantity,
                    'rate': item.rate,
                    'amount': item.line_total,
                    'order_total': order.grand_total,
                    'order_tax': order.tax_amount,
                    'order_discount': order.discount_amount,
                    'status': order.get_status_display(),
                    'status_code': order.status,
                    'authorization_status': auth_status,
                    'authorization_status_code': order.authorized_status or 0,
                    'country': country,
                    'state': state,
                    'state_id': state_id,
                    'district': district,
                    'mandal': mandal,
                    'city': city,
                    'city_id': city_id,
                    'area': area,
                    'area_id': area_id,
                    'agent_name': order.distributor.agent.name if order.distributor and order.distributor.agent else None,
                })
        
        return product_rows
    
    def list(self, request, *args, **kwargs):
        """Override list to return product-wise data with summary"""
        queryset = self.filter_queryset(self.get_queryset())
        
        summary = queryset.aggregate(
            total_orders=Count('id'),
            total_amount=Sum('grand_total'),
            total_tax=Sum('tax_amount'),
            total_product_count=Count('items'),
        )
        
        total_product_count = summary.pop('total_product_count') or 0
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            product_rows = self._convert_to_product_rows(page)
            serializer = self.get_serializer(product_rows, many=True)
            
            response = self.get_paginated_response(serializer.data)
            response.data['summary'] = {
                'total_orders': summary['total_orders'] or 0,
                'total_products': total_product_count,
                'total_amount': float(summary['total_amount'] or 0),
                'total_tax': float(summary['total_tax'] or 0),
            }
            return response
        
        product_rows = self._convert_to_product_rows(queryset)
        serializer = self.get_serializer(product_rows, many=True)
        return Response({
            'results': serializer.data,
            'count': len(product_rows),
            'summary': {
                'total_orders': summary['total_orders'] or 0,
                'total_products': len(product_rows),
                'total_amount': float(summary['total_amount'] or 0),
                'total_tax': float(summary['total_tax'] or 0),
            }
        })


class SalesOrderReportExportView(APIView):
    """
    Export Sales Order Report to Excel, CSV, or PDF
    """
    
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        export_format = request.data.get('format', 'excel').lower()
        filters_data = request.data.get('filters', {})
        
        if export_format not in ['excel', 'csv', 'pdf']:
            return Response(
                {'error': 'Invalid format. Must be excel, csv, or pdf'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self._get_filtered_queryset(request, filters_data)
        
        if export_format == 'excel':
            return self._export_excel(queryset)
        elif export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'pdf':
            return self._export_pdf(queryset)
    
    def _get_filtered_queryset(self, request, filters_data):
        """Apply filters to queryset"""
        queryset = SalesOrder.filtered_objects.get_qs(
            user=request.user
        ).select_related(
            'billing_state',
            'billing_city',
            'billing_area',
            'retailer',
            'distributor',
            'superstockist'
        ).prefetch_related(
            'items',
            'items__item'
        )
        
        queryset = apply_company_location_filter(queryset, request.user, company_field='company')
        filterset = SalesOrderReportFilter(filters_data, queryset=queryset, request=request)
        return filterset.qs
    
    def _export_excel(self, queryset):
        """Export to Excel format - Product-wise"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sales Order Report'
        
        headers = [
            'Order Number', 'Order Date', 'Customer Name', 'Customer Type',
            'Product Code', 'Product Name', 'Quantity', 'Rate', 'Amount',
            'Order Total', 'Status', 'Auth Status', 'State', 'City', 'Area',
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 2
        for order in queryset:
            state = order.billing_state.name if order.billing_state else ''
            city = order.billing_city.name if order.billing_city else ''
            area = order.billing_area.name if order.billing_area else ''
            
            auth_status_map = {1: 'Pending', 2: 'Approved', 3: 'Rejected'}
            auth_status = auth_status_map.get(order.authorized_status, 'Unknown')
            
            for item in order.items.all():
                row_data = [
                    order.order_number,
                    order.order_date.strftime('%Y-%m-%d'),
                    order.get_customer_name(),
                    order.get_customer_type_display(),
                    item.item.code,
                    item.item.name,
                    float(item.quantity),
                    float(item.rate),
                    float(item.line_total),
                    float(order.grand_total),
                    order.get_status_display(),
                    auth_status,
                    state,
                    city,
                    area,
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                
                row_num += 1
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sales_order_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response
    
    def _export_csv(self, queryset):
        """Export to CSV format - Product-wise"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=sales_order_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        writer = csv.writer(response)
        writer.writerow([
            'Order Number', 'Order Date', 'Customer Name', 'Customer Type',
            'Product Code', 'Product Name', 'Quantity', 'Rate', 'Amount',
            'Order Total', 'Status', 'Auth Status', 'State', 'City', 'Area',
        ])
        
        for order in queryset:
            state = order.billing_state.name if order.billing_state else ''
            city = order.billing_city.name if order.billing_city else ''
            area = order.billing_area.name if order.billing_area else ''
            
            auth_status_map = {1: 'Pending', 2: 'Approved', 3: 'Rejected'}
            auth_status = auth_status_map.get(order.authorized_status, 'Unknown')
            
            for item in order.items.all():
                writer.writerow([
                    order.order_number,
                    order.order_date.strftime('%Y-%m-%d'),
                    order.get_customer_name(),
                    order.get_customer_type_display(),
                    item.item.code,
                    item.item.name,
                    float(item.quantity),
                    float(item.rate),
                    float(item.line_total),
                    float(order.grand_total),
                    order.get_status_display(),
                    auth_status,
                    state,
                    city,
                    area,
                ])
        
        return response
    
    def _export_pdf(self, queryset):
        """Export to PDF format - Product-wise"""
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        
        title = Paragraph('Sales Order Report (Product-wise)', title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        data = [[
            'Order No.', 'Date', 'Customer', 'Product',
            'Qty', 'Rate', 'Amount', 'Status', 'Location',
        ]]
        
        row_count = 0
        for order in queryset:
            if row_count >= 100:
                break
                
            state = order.billing_state.name if order.billing_state else ''
            city = order.billing_city.name if order.billing_city else ''
            location = f"{state}, {city}" if city else state
            
            for item in order.items.all():
                if row_count >= 100:
                    break
                    
                data.append([
                    order.order_number,
                    order.order_date.strftime('%Y-%m-%d'),
                    order.get_customer_name()[:15],
                    item.item.name[:20],
                    f"{float(item.quantity):.2f}",
                    f"₹{float(item.rate):.2f}",
                    f"₹{float(item.line_total):.2f}",
                    order.get_status_display()[:10],
                    location[:20],
                ])
                row_count += 1
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=sales_order_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.write(pdf)
        return response
